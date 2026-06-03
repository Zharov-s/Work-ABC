import openpyxl, re, sys
from datetime import datetime
from urllib.parse import urlparse

sys.path.insert(0, 'webapp')
from database import get_db, sync_mailing_recipients

TODAY  = datetime.now().strftime('%Y-%m-%d')
MAX_PER_COMPANY = 5

EMAIL_RE = re.compile(r'[a-zA-Z0-9][a-zA-Z0-9._%+\-]*@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')
BLOCKED = {
    'info','sales','office','support','mail','contact','zakaz','hello','admin',
    'reception','corp','marketing','pr','press','media','hr','career',
    'communications','comms','post','inbox','noreply','no-reply','feedback',
    'help','service','request','quality','tender','zakupki','buh','director',
    'general','pressa','epd','infosecurity','ock','ekb','spb','dv','job',
    'anticorruption','vopros','ops','commercial','reklama','pismo','ib',
    'hotline','urgent','appeal','contact','zakaz','dms','gk',
}

def is_generic(email):
    local = email.split('@')[0].lower()
    if local in BLOCKED:
        return True
    for p in BLOCKED:
        if (local.startswith(p+'-') or local.startswith(p+'_') or
                local.startswith(p+'.') or
                (local.startswith(p) and len(local) > len(p) and local[len(p)].isdigit())):
            return True
    return False

def parse_emails(raw):
    if not raw:
        return [], []
    personal, generic, seen = [], [], set()
    for e in EMAIL_RE.findall(str(raw)):
        e = e.lower()
        if e in seen:
            continue
        seen.add(e)
        if is_generic(e):
            generic.append(e)
        else:
            personal.append(e)
    return personal, generic

PHONE_RE = re.compile(r'[\+]?[\d][\d\s\-\(\)]{6,18}[\d]')

def parse_phones(raw):
    if not raw:
        return None, None
    mob = gen = None
    for m in PHONE_RE.findall(str(raw)):
        d = re.sub(r'\D','',m)
        if len(d) < 7:
            continue
        is_mob = ((len(d)==11 and d[:2] in ('79','89')) or
                  (len(d)==10 and d[0]=='9'))
        if is_mob and not mob:
            mob = m.strip()
        elif not is_mob and not gen:
            gen = m.strip()
    return mob, gen

def normalize_url(raw):
    if not raw:
        return None
    s = str(raw).strip()
    if not s or s.lower() in ('none','null',''):
        return None
    if not s.startswith('http'):
        s = 'https://' + s
    try:
        p = urlparse(s)
        if p.netloc and '.' in p.netloc:
            return s.rstrip('/')
    except Exception:
        pass
    return None

def segment_from(code_raw, activity='', name=''):
    code = str(code_raw or '').strip()
    t    = (str(activity or '') + ' ' + str(name or '')).lower()

    if re.match(r'26\.[1-89]', code) or any(w in t for w in [
            'электрон','приборостроен','датчик','сенсор','полупровод','микроэлект']):
        return 'Электроника и приборостроение'
    if re.match(r'(21|32\.5)', code) or any(w in t for w in [
            'медицин','фармацевт','диагностик','хирург','стоматолог','биотех','медтех']):
        return 'Медтех и фармацевтика'
    if re.match(r'(62|63|58\.2|46\.5|95\.1|26\.2[01])', code) or any(w in t for w in [
            'информацион','программн','вычислит','компьютер','телеком','цифров',
            'it-','кибер','защитаинф','инфотранс']):
        return 'IT-производство и hardware'
    if re.match(r'(28|29|30|33)', code) or any(w in t for w in [
            'робот','автомат','станок','мехатрон','пневматик','гидравлик','серво']):
        return 'Робототехника и автоматизация'
    if re.match(r'26\.7', code) or any(w in t for w in [
            'лазер','оптик','фотон','световод','оптоэлектр']):
        return 'Лазерные и оптические технологии'
    if re.match(r'72\.', code):
        if any(w in t for w in ['медицин','биолог','фарм']):
            return 'Медтех и фармацевтика'
        if any(w in t for w in ['лазер','оптик']):
            return 'Лазерные и оптические технологии'
        if any(w in t for w in ['электрон','прибор']):
            return 'Электроника и приборостроение'
        return 'IT-производство и hardware'
    return 'Прочее light industrial'

def norm_name(n):
    n = str(n or '').strip().lower()
    n = re.sub(r'^(ооо|оао|зао|пао|ип|ао|нпо|нии|фгуп|мгуп|ргуп|муп|гуп|гбу|гку|гко)\s+','',n)
    n = re.sub(r'["""\'`]+', '', n)
    return n.strip()

# Load existing
conn = get_db()
existing_emails = {
    r['email'].lower()
    for r in conn.execute("SELECT email FROM contacts WHERE email IS NOT NULL").fetchall()
}
existing_companies = {
    norm_name(r['company_name'])
    for r in conn.execute("SELECT company_name FROM contacts WHERE company_name IS NOT NULL").fetchall()
    if r['company_name']
}
print('База: %d компаний, %d email' % (len(existing_companies), len(existing_emails)))

FILES = [
    ('Research/output/Новый сегмент 1.xlsx', 28),
    ('Research/output/Новый сегмент 2.xlsx', 16),
]

inserted = 0
skipped  = 0
company_counts = {}

for fpath, ncols in FILES:
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active
    print('\n%s: %d строк' % (fpath, ws.max_row - 1))

    for row_idx in range(2, ws.max_row + 1):
        def v(col):
            return ws.cell(row_idx, col).value

        if ncols == 28:
            company_name = v(1)
            inn          = str(v(2) or '').strip() or None
            director     = str(v(11) or '').strip() or None
            title        = str(v(13) or '').strip() or None
            phone_raw    = v(14)
            email_raw    = v(15)
            website_raw  = v(16)
            activity     = str(v(17) or '').strip()
            region_raw   = str(v(7) or 'г Москва').strip()
        else:
            company_name = v(1)
            inn          = str(v(2) or '').strip() or None
            director     = str(v(3) or '').strip() or None
            title        = str(v(5) or '').strip() or None
            phone_raw    = v(6)
            email_raw    = v(7)
            website_raw  = v(8)
            activity     = str(v(9) or '').strip()
            region_raw   = 'г Москва'

        if not company_name:
            continue

        company_name = str(company_name).strip()
        norm         = norm_name(company_name)
        website      = normalize_url(website_raw)
        mobile, gen_phone = parse_phones(phone_raw)
        personal_list, generic_list = parse_emails(email_raw)

        activity_code = activity[:7] if activity else ''
        segment = segment_from(activity_code, activity, company_name)

        region = 'Москва'
        rl = region_raw.lower()
        if 'московская' in rl or 'мо' == rl.strip():
            region = 'Московская область'
        elif 'москва' in rl:
            region = 'Москва'
        else:
            region = region_raw.replace('г ', '').replace('г.', '').strip() or 'Москва'

        existing_for_company = company_counts.get(norm, 0)
        rows_to_insert = []

        for pe in personal_list:
            if existing_for_company + len(rows_to_insert) >= MAX_PER_COMPANY:
                break
            if pe in existing_emails:
                continue
            rows_to_insert.append((pe, pe, None))

        for ge in generic_list:
            if existing_for_company + len(rows_to_insert) >= MAX_PER_COMPANY:
                break
            if ge in existing_emails:
                continue
            rows_to_insert.append((ge, None, ge))

        if not rows_to_insert:
            if not personal_list and not generic_list:
                if norm in existing_companies:
                    skipped += 1
                    continue
                rows_to_insert.append((None, None, None))
            else:
                skipped += len(personal_list) + len(generic_list)
                continue

        phone_val = mobile or gen_phone

        for (email_val, personal_email, generic_email) in rows_to_insert:
            try:
                conn.execute(
                    """INSERT OR IGNORE INTO contacts
                       (company_name, website, person_name, title,
                        email, personal_email, generic_email,
                        phone, mobile_phone, generic_phone,
                        inn, segment, region, date_found, status)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,'new')""",
                    (company_name, website, director, title,
                     email_val, personal_email, generic_email,
                     phone_val, mobile, gen_phone,
                     inn, segment, region, TODAY)
                )
                if conn.execute('SELECT changes()').fetchone()[0] > 0:
                    inserted += 1
                    if email_val:
                        existing_emails.add(email_val)
                    existing_companies.add(norm)
                    company_counts[norm] = existing_for_company + len(rows_to_insert)
                else:
                    skipped += 1
            except Exception as e:
                print('ERR row %d: %s' % (row_idx, e))

conn.commit()
print('\nСинхронизирую mailing_recipients...')
sync_mailing_recipients(conn)
conn.commit()
conn.close()

print('\nИмпорт завершён:')
print('  Добавлено строк: %d' % inserted)
print('  Пропущено:       %d' % skipped)

conn2 = get_db()
segs = conn2.execute(
    "SELECT segment, count(*) n FROM contacts GROUP BY segment ORDER BY n DESC"
).fetchall()
print('\nРаспределение по сегментам:')
for s in segs:
    print('  %-45s : %d' % (s['segment'] or '—', s['n']))
total = conn2.execute("SELECT count(*) FROM contacts").fetchone()[0]
print('\nВсего контактов в базе: %d' % total)
conn2.close()
