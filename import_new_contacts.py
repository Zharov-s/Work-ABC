import openpyxl, csv, re, sys, os
from datetime import datetime
from urllib.parse import urlparse

sys.path.insert(0, 'webapp')
from database import get_db, sync_mailing_recipients

TODAY  = datetime.now().strftime('%Y-%m-%d')
FOLDER = 'Research/output/Новые контакты'

# ── Email helpers ─────────────────────────────────────────────────────────────
EMAIL_RE = re.compile(r'[a-zA-Z0-9][a-zA-Z0-9._%+\-]*@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')
BLOCKED = {
    'info','sales','office','support','mail','contact','zakaz','hello','admin',
    'reception','corp','marketing','pr','press','media','hr','career',
    'communications','comms','post','inbox','noreply','no-reply','feedback',
    'help','service','request','quality','tender','zakupki','buh','director',
    'general','pressa','epd','infosecurity','ock','ekb','spb','dv','job',
    'anticorruption','vopros','ops','commercial','reklama','pismo','ib',
    'hotline','appeal','gk','dms','corp','owner','rent','daily',
    'sd','support','info','sales','office',
}

def is_generic(email):
    local = email.split('@')[0].lower()
    if local in BLOCKED: return True
    for p in BLOCKED:
        if (local.startswith(p+'-') or local.startswith(p+'_') or
                local.startswith(p+'.') or
                (local.startswith(p) and len(local)>len(p) and local[len(p)].isdigit())):
            return True
    return False

def extract_emails(raw):
    """Return (personal_list, generic_list)."""
    if not raw: return [],[]
    personal, generic, seen = [],[],set()
    for e in EMAIL_RE.findall(str(raw)):
        e=e.lower()
        if e in seen: continue
        seen.add(e)
        (generic if is_generic(e) else personal).append(e)
    return personal, generic

PHONE_RE = re.compile(r'[\+]?[\d][\d\s\-\(\)\.]{6,20}[\d]')

def classify_phone(raw):
    """Return (mobile, generic) for a single phone string."""
    if not raw: return None, None
    d = re.sub(r'\D','',str(raw))
    if len(d) < 7: return None, None
    is_mob = ((len(d)==11 and d[:2] in ('79','89')) or
              (len(d)==10 and d[0]=='9'))
    phone = str(raw).strip()
    return (phone,None) if is_mob else (None,phone)

def normalize_url(raw):
    if not raw: return None
    s = str(raw).strip()
    if not s or s.lower() in ('none','null',''): return None
    if not s.startswith('http'): s='https://'+s
    try:
        p=urlparse(s)
        if p.netloc and '.' in p.netloc:
            return s.rstrip('/')
    except: pass
    return None

# ── Segmentation (name/website keywords — no ОКВЭД here) ─────────────────────
def segment_from_name(name, website=''):
    t = (str(name or '')+' '+str(website or '')).lower()

    if any(w in t for w in ['электрон','электр','прибор','датчик','сенсор','полупровод',
                              'микро','rvi','ied','bms','led','лед','свет','энерджи',
                              'энерго','power','lighting','luminis','rkor','riscom']):
        return 'Электроника и приборостроение'
    if any(w in t for w in ['медтех','medtec','medtech','мед ','медицин','фармацевт',
                              'диагностик','хирург','стоматолог','meko','меко']):
        return 'Медтех и фармацевтика'
    if any(w in t for w in ['автоматик','автомат','робот','станок','мехатрон','пневмат',
                              'гидравл','серво','плк','plc','mts','insist','инсист',
                              'systems','смарт','smart','grade','beward','бевард']):
        return 'Робототехника и автоматизация'
    if any(w in t for w in ['информацион','программ','вычислит','компьютер','телеком',
                              'цифров','it-','кибер','защита','инфо','cyber','digital',
                              'ntc','нтц','нпц','сервер','cloud','ит ']):
        return 'IT-производство и hardware'
    if any(w in t for w in ['лазер','оптик','фотон','световод','оптоэлектр']):
        return 'Лазерные и оптические технологии'
    return 'Прочее light industrial'

def norm_name(n):
    n=str(n or '').strip().lower()
    n=re.sub(r'^(ооо|оао|зао|пао|ип|ао|нпо|нии|фгуп|мгуп|ргуп|муп|гуп|гбу|гку|гко|ад)\s+','',n)
    n=re.sub(r'["""\'`]+','',n)
    return n.strip()

# ── Load existing DB ──────────────────────────────────────────────────────────
conn = get_db()
existing_emails = {
    r['email'].lower()
    for r in conn.execute("SELECT email FROM contacts WHERE email IS NOT NULL").fetchall()
}
existing_companies = {
    norm_name(r['company_name'])
    for r in conn.execute("SELECT company_name FROM contacts WHERE company_name IS NOT NULL").fetchall()
    if r['company_name']
}
print('Перед импортом: %d компаний, %d email в базе' % (len(existing_companies),len(existing_emails)))

# ── Generic INSERT ─────────────────────────────────────────────────────────────
INSERT_SQL = """INSERT OR IGNORE INTO contacts
  (company_name,website,person_name,title,
   email,personal_email,generic_email,
   phone,mobile_phone,generic_phone,
   segment,region,date_found,status)
  VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,'new')"""

inserted = 0
skipped  = 0

def insert_row(company_name, website, person, title, email_val,
               personal_email, generic_email, mobile, gen_phone, segment):
    global inserted, skipped
    norm = norm_name(company_name)
    if email_val and email_val in existing_emails:
        skipped += 1
        return False
    phone_val = mobile or gen_phone
    conn.execute(INSERT_SQL, (
        company_name, website, person or None, title or None,
        email_val or None, personal_email or None, generic_email or None,
        phone_val or None, mobile or None, gen_phone or None,
        segment, 'Москва', TODAY
    ))
    if conn.execute('SELECT changes()').fetchone()[0] > 0:
        inserted += 1
        if email_val:
            existing_emails.add(email_val)
        existing_companies.add(norm)
        return True
    else:
        skipped += 1
        return False

# ── Process standard XLSX files ───────────────────────────────────────────────
xlsx_files = sorted([f for f in os.listdir(FOLDER)
                     if f.endswith('.xlsx') and not f.startswith('.')])

for fname in xlsx_files:
    path = os.path.join(FOLDER, fname)
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        file_inserted = 0
        for row_idx in range(2, ws.max_row+1):
            def v(c): return ws.cell(row_idx,c).value
            company_name = v(1)
            if not company_name: continue
            company_name = str(company_name).strip()
            website = normalize_url(v(2))
            person  = str(v(3) or '').strip() or None
            title   = str(v(4) or '').strip() or None
            email_raw = v(5)
            phone_raw = str(v(6) or '').strip()
            segment   = segment_from_name(company_name, str(v(2) or ''))

            # Skip if 8th+ column has "не актуально"
            note = ''
            if ws.max_column >= 8:
                note = str(v(8) or '').strip().lower()

            personal_list, generic_list = extract_emails(email_raw)
            mobile, gen_phone = classify_phone(phone_raw)

            # Insert one row per unique email
            if personal_list or generic_list:
                for pe in personal_list:
                    insert_row(company_name,website,person,title,pe,pe,None,mobile,gen_phone,segment)
                    file_inserted += 1
                for ge in generic_list:
                    insert_row(company_name,website,person,title,ge,None,ge,mobile,gen_phone,segment)
                    file_inserted += 1
            else:
                # No email — still save if company not in DB
                if norm_name(company_name) not in existing_companies:
                    insert_row(company_name,website,person,title,None,None,None,mobile,gen_phone,segment)
                    file_inserted += 1

        wb.close()
        if file_inserted > 0:
            print('  %s: +%d' % (fname, file_inserted))
    except Exception as e:
        print('  %s: ERROR %s' % (fname, e))

# ── Process new_25_contacts16.csv ─────────────────────────────────────────────
csv_path = os.path.join(FOLDER, 'new_25_contacts16.csv')
try:
    with open(csv_path, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames
        file_inserted = 0
        for row in reader:
            company_name = (row.get('Наименование') or '').strip()
            if not company_name: continue
            website  = normalize_url(row.get('Сайт'))
            person   = (row.get('ФИО') or '').strip() or None
            title    = (row.get('Должность') or '').strip() or None
            email_raw= row.get('Емайл') or row.get('Email') or ''
            phone_raw= (row.get('Телефон') or '').strip()
            segment  = segment_from_name(company_name, str(website or ''))

            personal_list, generic_list = extract_emails(email_raw)
            mobile, gen_phone = classify_phone(phone_raw)
            if personal_list or generic_list:
                for pe in personal_list:
                    insert_row(company_name,website,person,title,pe,pe,None,mobile,gen_phone,segment)
                    file_inserted += 1
                for ge in generic_list:
                    insert_row(company_name,website,person,title,ge,None,ge,mobile,gen_phone,segment)
                    file_inserted += 1
            else:
                if norm_name(company_name) not in existing_companies:
                    insert_row(company_name,website,person,title,None,None,None,mobile,gen_phone,segment)
                    file_inserted += 1
        if file_inserted > 0:
            print('  new_25_contacts16.csv: +%d' % file_inserted)
except Exception as e:
    print('  new_25_contacts16.csv: ERROR %s' % e)

# ── Process mitinskaya_16_yes_contacts_audit.csv (UTF-16, tab-sep) ────────────
mit_path = os.path.join(FOLDER, 'mitinskaya_16_yes_contacts_audit.csv')
try:
    with open(mit_path, encoding='utf-16') as f:
        content = f.read()
    rows_raw = [l.split('\t') for l in content.strip().split('\n')]
    file_inserted = 0
    # Col indices: 1=Company, 5=Site, 6=LPR, 7=Title, 8=DirectEmail, 9=DirectMobile, 10=AllEmails, 11=OrgMobile, 12=OrgPhone
    for r in rows_raw[1:]:
        r = r + ['']*15
        company_name = r[1].strip()
        if not company_name or company_name == 'Компания / организация': continue
        website  = normalize_url(r[5].strip())
        person   = r[6].strip() or None
        title    = r[7].strip() or None
        segment  = segment_from_name(company_name, str(website or ''))

        # Collect ALL emails from both fields (direct + all found)
        all_email_raw = r[8] + '; ' + r[10]
        personal_list, generic_list = extract_emails(all_email_raw.replace(';',' ').replace(',',' '))

        # Phones
        mobile_raw   = (r[9] + ' ' + r[11]).strip()
        gen_phone_raw = r[12].strip()
        mobile, _    = classify_phone(mobile_raw) if mobile_raw.strip() else (None,None)
        _, gen_phone = classify_phone(gen_phone_raw) if gen_phone_raw.strip() else (None,None)

        if personal_list or generic_list:
            for pe in personal_list:
                insert_row(company_name,website,person,title,pe,pe,None,mobile,gen_phone,segment)
                file_inserted += 1
            for ge in generic_list:
                insert_row(company_name,website,person,title,ge,None,ge,mobile,gen_phone,segment)
                file_inserted += 1
        else:
            if norm_name(company_name) not in existing_companies:
                insert_row(company_name,website,person,title,None,None,None,mobile,gen_phone,segment)
                file_inserted += 1
    if file_inserted > 0:
        print('  mitinskaya_16_yes_contacts_audit.csv: +%d' % file_inserted)
except Exception as e:
    print('  mitinskaya_16_yes_contacts_audit.csv: ERROR %s' % e)

conn.commit()
print('\nВставлено: %d | Пропущено: %d' % (inserted, skipped))

# ══════════════════════════════════════════════════════════════════════════════
# ДЕДУПЛИКАЦИЯ
# ══════════════════════════════════════════════════════════════════════════════
print('\n── Дедупликация ──')

# 1. Найти строки с NULL email где та же компания имеет строку с email
dup1 = conn.execute("""
    SELECT c.id
    FROM contacts c
    WHERE c.email IS NULL
    AND EXISTS (
        SELECT 1 FROM contacts c2
        WHERE lower(c2.company_name) = lower(c.company_name)
        AND c2.email IS NOT NULL
        AND c2.id != c.id
    )
""").fetchall()
ids1 = [r['id'] for r in dup1]
if ids1:
    conn.execute('DELETE FROM contacts WHERE id IN (%s)' % ','.join('?'*len(ids1)), ids1)
    print('  Удалено пустых email-строк (компания есть с email): %d' % len(ids1))

# 2. Полные дубликаты по email (одинаковый email — оставить строку с меньшим id)
dup2 = conn.execute("""
    SELECT id FROM contacts
    WHERE email IS NOT NULL
    AND id NOT IN (
        SELECT MIN(id) FROM contacts
        WHERE email IS NOT NULL
        GROUP BY lower(email)
    )
""").fetchall()
ids2 = [r['id'] for r in dup2]
if ids2:
    conn.execute('DELETE FROM contacts WHERE id IN (%s)' % ','.join('?'*len(ids2)), ids2)
    print('  Удалено дублей по email: %d' % len(ids2))

# 3. Дубликаты по (company_name + person_name + NULL email) — оставить одну
dup3 = conn.execute("""
    SELECT id FROM contacts
    WHERE email IS NULL AND person_name IS NOT NULL
    AND id NOT IN (
        SELECT MIN(id) FROM contacts
        WHERE email IS NULL AND person_name IS NOT NULL
        GROUP BY lower(company_name), lower(person_name)
    )
""").fetchall()
ids3 = [r['id'] for r in dup3]
if ids3:
    conn.execute('DELETE FROM contacts WHERE id IN (%s)' % ','.join('?'*len(ids3)), ids3)
    print('  Удалено дублей (компания+ФИО без email): %d' % len(ids3))

if not ids1 and not ids2 and not ids3:
    print('  Дублей не найдено')

conn.commit()

# ── Sync mailing recipients ───────────────────────────────────────────────────
print('\nСинхронизирую mailing_recipients...')
sync_mailing_recipients(conn)
conn.commit()
conn.close()

# ── Final stats ───────────────────────────────────────────────────────────────
conn2 = get_db()
total = conn2.execute('SELECT count(*) FROM contacts').fetchone()[0]
mailing = conn2.execute('SELECT count(*) FROM mailing_recipients').fetchone()[0]
segs = conn2.execute('SELECT segment, count(*) n FROM contacts GROUP BY segment ORDER BY n DESC').fetchall()

print('\n══ ИТОГ ══')
print('Всего контактов: %d' % total)
print('В очереди рассылки: %d' % mailing)
print('\nПо сегментам:')
for s in segs:
    print('  %-45s : %d' % (s['segment'] or '—', s['n']))
conn2.close()
