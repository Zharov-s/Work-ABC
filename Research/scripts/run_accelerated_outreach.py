from __future__ import annotations

import argparse
import csv
import json
import subprocess
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Sequence
from urllib.parse import urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "output" / "new_contacts.xlsx"
DEFAULT_MEMORY = ROOT / "output" / "found_contacts_memory.jsonl"
DEFAULT_EVIDENCE = ROOT / "output" / "evidence_log.csv"
DEFAULT_PIPELINE_REPORT = ROOT / "output" / "last_accelerated_pipeline.json"
DEFAULT_SEND_REPORT = ROOT / "output" / "last_email_send_report.json"
DEFAULT_COPY_EMAIL = "s.zharov@abcentrum.ru"
DEFAULT_MAX_TOTAL_RECIPIENTS = 29
DEFAULT_EXPECTED_ROWS = 25
EXPECTED_SHEET = "new_contacts"
EXPECTED_HEADERS = [
    "Наименование",
    "Сайт",
    "ФИО",
    "Должность",
    "Емайл",
    "Мобильный телефон или городской с добавочным",
]


def normalize(value: object) -> str:
    return (value or "").strip() if isinstance(value, str) else str(value or "").strip()


def normalize_company(value: object) -> str:
    return " ".join(normalize(value).lower().split())


def normalize_website(value: object) -> str:
    raw = normalize(value).lower()
    if not raw:
        return ""
    candidate = raw if "://" in raw else f"https://{raw}"
    parsed = urlparse(candidate)
    host = (parsed.netloc or parsed.path).lower().strip().strip("/")
    if host.startswith("www."):
        host = host[4:]
    path = parsed.path.strip("/") if parsed.netloc else ""
    return f"{host}/{path}" if path else host


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Prepare or send an accelerated outreach batch from output/new_contacts.xlsx."
    )
    parser.add_argument(
        "--stage",
        choices=("prepare", "send", "full"),
        default="prepare",
        help=(
            "prepare: format, validate, append to memory, and stop for user confirmation; "
            "send: send the already prepared workbook; "
            "full: run both stages in one command."
        ),
    )
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--memory", default=str(DEFAULT_MEMORY))
    parser.add_argument("--evidence", default=str(DEFAULT_EVIDENCE))
    parser.add_argument("--pipeline-report", default=str(DEFAULT_PIPELINE_REPORT))
    parser.add_argument("--send-report", default=str(DEFAULT_SEND_REPORT))
    parser.add_argument(
        "--expected-rows",
        type=int,
        default=DEFAULT_EXPECTED_ROWS,
        help="Expected number of filled rows in output/new_contacts.xlsx.",
    )
    parser.add_argument("--copy-email", default=DEFAULT_COPY_EMAIL)
    parser.add_argument(
        "--max-total-recipients",
        type=int,
        default=DEFAULT_MAX_TOTAL_RECIPIENTS,
        help="Maximum total recipients per email, including the required copy email.",
    )
    parser.add_argument("--subject", help="Optional subject override for the email batches.")
    parser.add_argument(
        "--dry-run-email",
        action="store_true",
        help="Run the email stage in dry-run mode without sending.",
    )
    parser.add_argument(
        "--skip-validation",
        action="store_true",
        help="Skip strict validation. Use only for explicit resend/debug scenarios.",
    )
    parser.add_argument(
        "--skip-memory-append",
        action="store_true",
        help="Skip appending to found_contacts_memory.jsonl. Useful for resends.",
    )
    parser.add_argument("--start-batch", type=int, default=1)
    parser.add_argument("--end-batch", type=int)
    return parser.parse_args(argv)


def run_command(command: list[str], *, cwd: Path) -> str:
    result = subprocess.run(
        command,
        cwd=cwd,
        capture_output=True,
        text=True,
        check=False,
    )
    output = "\n".join(part for part in [result.stdout.strip(), result.stderr.strip()] if part).strip()
    if result.returncode != 0:
        raise RuntimeError(output or f"Command failed: {' '.join(command)}")
    return output


def load_batch_rows(workbook_path: Path) -> list[dict[str, str]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if EXPECTED_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet `{EXPECTED_SHEET}` in {workbook_path}")
    ws = wb[EXPECTED_SHEET]
    headers = [normalize(cell.value) for cell in ws[1][: len(EXPECTED_HEADERS)]]
    if headers != EXPECTED_HEADERS:
        raise RuntimeError(f"Header mismatch in {workbook_path}: {headers}")

    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS), values_only=True):
        values = [normalize(value) for value in row]
        if not any(values):
            continue
        rows.append(
            {
                "company_name": values[0],
                "website": values[1],
                "person_name": values[2],
                "title": values[3],
                "email": values[4],
                "phone": values[5],
            }
        )
    return rows


def load_memory_keys(memory_path: Path) -> tuple[set[str], set[str]]:
    companies: set[str] = set()
    websites: set[str] = set()
    if not memory_path.exists():
        return companies, websites

    for raw_line in memory_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        try:
            record = json.loads(line)
        except json.JSONDecodeError:
            continue
        companies.add(normalize_company(record.get("company_name", "")))
        websites.add(normalize_website(record.get("website", "")))
    return companies, websites


def build_source_lookup(evidence_path: Path) -> dict[str, dict[object, str]]:
    lookup: dict[str, dict[object, str]] = {
        "company": {},
        "company_person": {},
        "company_person_contact": {},
    }
    if not evidence_path.exists():
        return lookup

    with evidence_path.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            source_url = normalize(row.get("source_url", ""))
            company = normalize_company(row.get("legal_entity", ""))
            person = normalize(row.get("person_name", "")).lower()
            contact_value = normalize(row.get("contact_value", "")).lower()
            if not source_url or not company:
                continue
            lookup["company"][company] = source_url
            if person:
                lookup["company_person"][(company, person)] = source_url
            if person and contact_value:
                lookup["company_person_contact"][(company, person, contact_value)] = source_url
    return lookup


def resolve_source_url(
    row: dict[str, str],
    lookup: dict[str, dict[object, str]],
) -> str:
    company = normalize_company(row["company_name"])
    person = normalize(row["person_name"]).lower()
    email = normalize(row["email"]).lower()
    phone = normalize(row["phone"]).lower()

    for key in (
        (company, person, email),
        (company, person, phone),
    ):
        value = lookup["company_person_contact"].get(key)
        if value:
            return value

    for key in (
        (company, person),
        company,
    ):
        if isinstance(key, tuple):
            value = lookup["company_person"].get(key)
        else:
            value = lookup["company"].get(key)
        if value:
            return value
    return ""


def append_memory_records(
    *,
    rows: list[dict[str, str]],
    workbook_path: Path,
    memory_path: Path,
    evidence_path: Path,
) -> list[dict[str, str]]:
    existing_companies, existing_websites = load_memory_keys(memory_path)
    source_lookup = build_source_lookup(evidence_path)
    batch_file = str(workbook_path.relative_to(ROOT)) if workbook_path.is_relative_to(ROOT) else str(workbook_path)
    batch_type = f"new_contacts_exact_{len(rows)}"
    appended: list[dict[str, str]] = []

    memory_path.parent.mkdir(parents=True, exist_ok=True)
    with memory_path.open("a", encoding="utf-8") as f:
        for row in rows:
            company_key = normalize_company(row["company_name"])
            website_key = normalize_website(row["website"])
            if company_key in existing_companies or website_key in existing_websites:
                raise RuntimeError(
                    f"Refusing to append duplicate company to memory: {row['company_name']} / {row['website']}"
                )

            record = {
                "date_found": date.today().isoformat(),
                "batch_file": batch_file,
                "company_name": row["company_name"],
                "website": row["website"],
                "person_name": row["person_name"],
                "title": row["title"],
                "email": row["email"],
                "phone": row["phone"],
                "source_url": resolve_source_url(row, source_lookup),
                "batch_type": batch_type,
            }
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
            appended.append(record)
            existing_companies.add(company_key)
            existing_websites.add(website_key)

    return appended


def main(argv: Sequence[str] | None = None) -> None:
    args = parse_args(argv)
    workbook_path = Path(args.workbook).resolve()
    memory_path = Path(args.memory).resolve()
    evidence_path = Path(args.evidence).resolve()
    pipeline_report_path = Path(args.pipeline_report).resolve()
    send_report_path = Path(args.send_report).resolve()
    should_prepare = args.stage in {"prepare", "full"}
    should_send = args.stage in {"send", "full"}

    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    format_output = "format skipped"
    if should_prepare:
        format_output = run_command(
            [
                sys.executable,
                str(ROOT / "scripts" / "format_new_contacts.py"),
                "--workbook",
                str(workbook_path),
            ],
            cwd=ROOT,
        )

    validation_output = "validation skipped"
    if not args.skip_validation:
        validation_command = [
            sys.executable,
            str(ROOT / "validation" / "validate_new_contacts.py"),
            "--workbook",
            str(workbook_path),
            "--expected-rows",
            str(args.expected_rows),
        ]
        if args.stage == "send":
            validation_command.append("--skip-memory-check")
        validation_output = run_command(
            validation_command,
            cwd=ROOT,
        )

    batch_rows = load_batch_rows(workbook_path)
    if len(batch_rows) != args.expected_rows:
        raise RuntimeError(
            f"Workbook must contain exactly {args.expected_rows} filled rows; found {len(batch_rows)}"
        )

    memory_records: list[dict[str, str]] = []
    if should_prepare and not args.skip_memory_append:
        memory_records = append_memory_records(
            rows=batch_rows,
            workbook_path=workbook_path,
            memory_path=memory_path,
            evidence_path=evidence_path,
        )

    send_output = "email stage skipped"
    if should_send:
        send_command = [
            sys.executable,
            str(ROOT / "Pro-email" / "scripts" / "send_new_contacts_batches.py"),
            "--workbook",
            str(workbook_path),
            "--report",
            str(send_report_path),
            "--copy-email",
            args.copy_email,
            "--max-total-recipients",
            str(args.max_total_recipients),
            "--start-batch",
            str(args.start_batch),
        ]
        if args.end_batch is not None:
            send_command.extend(["--end-batch", str(args.end_batch)])
        if args.subject:
            send_command.extend(["--subject", args.subject])
        if args.dry_run_email:
            send_command.append("--dry-run")

        send_output = run_command(send_command, cwd=ROOT)

    pipeline_report = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "stage": args.stage,
        "workbook": str(workbook_path),
        "memory_path": str(memory_path),
        "send_report": str(send_report_path),
        "expected_rows": args.expected_rows,
        "batch_row_count": len(batch_rows),
        "copy_email": args.copy_email,
        "max_total_recipients": args.max_total_recipients,
        "dry_run_email": args.dry_run_email,
        "validation_ran": not args.skip_validation,
        "memory_appended": should_prepare and not args.skip_memory_append,
        "memory_appended_count": len(memory_records),
        "ready_for_send": args.stage == "prepare",
        "send_confirmation_required": args.stage == "prepare",
        "email_stage_completed": should_send,
        "format_output": format_output,
        "validation_output": validation_output,
        "send_output": send_output,
    }
    pipeline_report_path.parent.mkdir(parents=True, exist_ok=True)
    pipeline_report_path.write_text(
        json.dumps(pipeline_report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    if args.stage == "prepare":
        print(
            f"Accelerated batch prepared for {len(batch_rows)} row(s). "
            f"Memory appended: {len(memory_records)} row(s). "
            f"Ready for user confirmation before sending. "
            f"Pipeline report saved to {pipeline_report_path}"
        )
    elif args.stage == "send":
        print(
            f"Accelerated email send complete for {len(batch_rows)} row(s). "
            f"Pipeline report saved to {pipeline_report_path}"
        )
    else:
        print(
            f"Accelerated prepare+send pipeline complete for {len(batch_rows)} row(s). "
            f"Memory appended: {len(memory_records)} row(s). "
            f"Pipeline report saved to {pipeline_report_path}"
        )


if __name__ == "__main__":
    main()
