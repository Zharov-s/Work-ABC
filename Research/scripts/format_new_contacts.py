from __future__ import annotations

import argparse
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "output" / "new_contacts.xlsx"
SHEET_NAME = "new_contacts"
COLUMN_WIDTH = 30


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Apply the required accelerated formatting to output/new_contacts.xlsx."
    )
    parser.add_argument("--workbook", default=str(WORKBOOK))
    parser.add_argument("--sheet", default=SHEET_NAME)
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> None:
    args = parse_args(argv)
    workbook_path = Path(args.workbook).resolve()
    wb = load_workbook(workbook_path)
    ws = wb[args.sheet]

    for idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTH

    alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = alignment

    wb.save(workbook_path)


if __name__ == "__main__":
    main()
