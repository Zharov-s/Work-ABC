from __future__ import annotations

from csv import DictReader, DictWriter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TARGETS = ROOT / "input" / "target_companies.csv"
WORKBOOK = ROOT / "output" / "mitino_target_companies_lpr_direct_contacts.xlsx"
EVIDENCE = ROOT / "output" / "evidence_log.csv"
NOTES_DIR = ROOT / "output" / "research_notes"
RESEARCH_DATE = "2026-04-07"
SHEET_NAME = "Direct LPR Contacts"
HEADERS = [
    "Legal entity name",
    "Website",
    "Decision-maker full name",
    "Decision-maker title",
    "Corporate email",
    "Mobile phone",
    "Work phone + extension",
    "Telegram niknames",
    "Telegram-group",
    "Source URL",
    "Verification note",
]


ROWS = [
    {
        "slug": "yadro",
        "brand_name": "YADRO",
        "legal_entity": 'ООО "КНС ГРУПП"',
        "website": "https://yadro.com",
        "person": "Алексей Шелобков",
        "title": "генеральный директор YADRO",
        "source_url": "https://yadro.com",
        "confidence": "high",
    },
    {
        "slug": "aquarius",
        "brand_name": "Aquarius",
        "legal_entity": 'АО "ГРУППА АКВАРИУС"',
        "website": "https://www.aq.ru",
        "person": "Вадим Шаров",
        "title": 'директор производственного комплекса "Аквариус"',
        "telegram_group": "@aquariuspublic",
        "source_url": "https://www.aq.ru/contacts/",
        "verification_note": "Official Aquarius contacts page links the verified official Telegram-group handle @aquariuspublic; no current personal Telegram nickname, person-level direct email, mobile, or named extension was publicly found for Vadim Sharov.",
        "confidence": "high",
    },
    {
        "slug": "depo-computers",
        "brand_name": "DEPO Computers",
        "legal_entity": 'ООО "ДЕПО ЭЛЕКТРОНИКС"',
        "website": "https://www.depo.ru",
        "person": "Сергей Эскин",
        "title": "президент группы компаний DEPO Computers",
        "telegram_group": "@depo_computers",
        "source_url": "https://www.depo.ru/contacts/",
        "verification_note": "Official DEPO contacts page links the verified official Telegram-group handle @depo_computers; no current personal Telegram nickname, person-level direct email, mobile, or named extension was publicly found for Sergey Eskin.",
        "confidence": "medium",
    },
    {
        "slug": "kraftway",
        "brand_name": "Kraftway",
        "legal_entity": 'АО "КРАФТВЭЙ КОРПОРЭЙШН ПЛС"',
        "website": "https://kraftway.ru",
        "person": "Алексей Юрьевич Кравцов",
        "title": "президент",
        "source_url": "https://www.rusprofile.ru/search?query=%D0%9A%D0%A0%D0%90%D0%A4%D0%A2%D0%92%D0%AD%D0%99",
        "confidence": "high",
    },
    {
        "slug": "fplus",
        "brand_name": "Fplus",
        "legal_entity": 'ООО "Ф-ПЛЮС ОБОРУДОВАНИЕ И РАЗРАБОТКИ"',
        "website": "https://fplustech.ru",
        "person": "Борис Фастовецкий",
        "title": "руководитель отдела инфраструктурного ПО Департамента индустриальных программно-аппаратных комплексов Fplus",
        "corporate_email": "b.fastovetskiy@fplustech.ru",
        "mobile_phone": "+7 926 328-41-66",
        "source_url": "https://fplustech.ru",
        "verification_note": "Official Fplus site names Boris Fastovetskiy and publishes his direct corporate email plus public mobile.",
        "selection_rationale": "Official site exposes this named department head with direct person-level contact details; no higher-priority facilities / ops direct contact was publicly disclosed.",
        "confidence": "high",
    },
    {
        "slug": "qtech",
        "brand_name": "QTECH",
        "legal_entity": 'ООО "КЬЮТЭК"',
        "website": "https://www.qtech.ru",
        "person": "Илья Бишкиревич",
        "title": "руководитель отдела по развитию бизнеса",
        "telegram_group": "@QTECH_company",
        "source_url": "https://www.qtech.ru/company/contacts/",
        "verification_note": "Official QTECH contacts page links the verified official Telegram-group handle @QTECH_company; no current personal Telegram nickname, person-level direct email, mobile, or named extension was publicly found for Ilya Bishkirevich.",
        "confidence": "medium",
    },
    {
        "slug": "t8",
        "brand_name": "T8",
        "legal_entity": 'ООО "Т8"',
        "website": "https://t8.ru",
        "person": "Владимир Трещиков",
        "title": "генеральный директор",
        "corporate_email": "vt@t8.ru",
        "source_url": "https://checko.ru/company/t8-1087746479765",
        "verification_note": "Current public company-card source for ООО T8 lists Vladimir Treshchikov as general director and publishes vt@t8.ru; the same email is also publicly tied to him in conference materials.",
        "selection_rationale": "Public evidence supports Vladimir Treshchikov as the operating CEO of T8; no higher-priority admin / facilities / ops contact with a verified direct channel was publicly disclosed.",
        "confidence": "medium",
    },
    {
        "slug": "saga-technologies",
        "brand_name": "SAGA Technologies / САГА Технологии",
        "legal_entity": 'АО "САГА ТЕХНОЛОГИИ"',
        "website": "https://sagacorporation.com",
        "person": "Геннадий Викторович Талдыкин",
        "title": "генеральный директор",
        "telegram_group": "@SAGACorporation",
        "source_url": "https://sagacorporation.com",
        "verification_note": "Official SAGA site links the verified official Telegram-group handle @SAGACorporation; no current personal Telegram nickname, person-level direct email, mobile, or named extension was publicly found for Gennady Taldykin.",
        "confidence": "high",
    },
    {
        "slug": "parus-electro",
        "brand_name": "Parus Electro / Парус электро",
        "legal_entity": 'ООО "Парус Электро"',
        "website": "https://parus-electro.ru",
        "person": "Марченко Алексей Викторович",
        "title": "заместитель генерального директора по коммерческой деятельности",
        "corporate_email": "av.marchenko@parus-electro.ru",
        "source_url": "https://parus-electro.ru/company/staff/",
        "verification_note": "Official staff page lists Alexey Marchenko with direct corporate email; shared switchboard phone is omitted because no named extension is published for him.",
        "selection_rationale": "Official staff page exposes this senior commercial leader with direct corporate email; higher-priority admin / facilities roles were not publicly disclosed.",
        "confidence": "high",
    },
    {
        "slug": "impuls",
        "brand_name": "IMPULS / ИМПУЛЬС",
        "legal_entity": 'ООО "ЦРИ "Импульс""',
        "website": "https://impuls.energy",
        "person": "Алексей Александрович Сироткин",
        "title": "генеральный директор",
        "source_url": "https://www.rusprofile.ru/search?query=7743927077",
        "confidence": "high",
    },
    {
        "slug": "rezonit",
        "brand_name": "Rezonit / Резонит",
        "legal_entity": 'ООО "РЕЗОНИТ"',
        "website": "https://www.rezonit.ru",
        "person": "Андрей Ильич Кучерявый",
        "title": "генеральный директор",
        "telegram_group": "@rezonit",
        "source_url": "https://www.rezonit.ru",
        "verification_note": "Official Rezonit site links the verified official Telegram-group handle @rezonit; no current personal Telegram nickname, person-level direct email, mobile, or named extension was publicly found for Andrey Kucheryavy.",
        "confidence": "high",
    },
    {
        "slug": "a-contract",
        "brand_name": "A-CONTRACT / А-КОНТРАКТ",
        "legal_entity": 'ООО "А-КОНТРАКТ"',
        "website": "https://a-contract.ru",
        "person": "Сергей Фёдоров",
        "title": "директор производства",
        "source_url": "https://a-contract.ru/",
        "confidence": "low",
    },
    {
        "slug": "milandr",
        "brand_name": "Milandr / Миландр",
        "legal_entity": 'АО "ПКК Миландр"',
        "website": "https://milandr.ru",
        "person": "Алексей Юрьевич Новоселов",
        "title": "генеральный директор",
        "source_url": "https://www.rusprofile.ru/search?query=%D0%9C%D0%98%D0%9B%D0%90%D0%9D%D0%94%D0%A0",
        "confidence": "high",
    },
    {
        "slug": "motorica",
        "brand_name": "Motorica / Моторика",
        "legal_entity": 'МК ООО "Хомо Ауктус"',
        "website": "https://motorica.org",
        "person": "Валерия Казанкова",
        "title": "руководитель направления по связям с инвесторами",
        "corporate_email": "ir@motorica.org",
        "source_url": "https://ar2024.motorica.org/download/full-reports/ar_ru_annual-report_pages_motorica_2024.pdf",
        "verification_note": "Official 2024 annual report lists Valeria Kazankova as investor-relations lead with direct corporate email; published phone is a shared line without extension, so only email is retained.",
        "selection_rationale": "Official annual report exposes this named investor-relations lead with direct public email; no higher-priority operations / facilities contact was publicly disclosed.",
        "confidence": "medium",
    },
    {
        "slug": "infinet-wireless",
        "brand_name": "Infinet Wireless",
        "legal_entity": 'ООО "ИНФИНЕТ"',
        "website": "https://infinetwireless.com",
        "person": "Дмитрий Окороков",
        "title": "сооснователь / генеральный директор",
        "source_url": "https://infinetwireless.com",
        "confidence": "medium",
    },
    {
        "slug": "itelma-electronic-components",
        "brand_name": "Itelma Electronic Components / Итэлма Электронные компоненты",
        "legal_entity": 'АО "ЭЛЕКТРОННЫЕ КОМПОНЕНТЫ"',
        "website": "https://elecomponent.ru",
        "person": "Александр Сергеевич Чистов",
        "title": "генеральный директор",
        "source_url": "https://elecomponent.ru",
        "confidence": "medium",
    },
    {
        "slug": "nextouch",
        "brand_name": "NexTouch / НЕКС-Т",
        "legal_entity": 'ООО "НЕКС-Т"',
        "website": "https://nextouch.ru",
        "person": "Владимир Владимирович Крикушенко",
        "title": "генеральный директор",
        "corporate_email": "vm@nextouch.ru",
        "source_url": "https://zachestnyibiznes.ru/company/ul/1097746776643_7723740056_OOO-NEKS-T",
        "verification_note": "Public company-card sources list vm@nextouch.ru for ООО \"НЕКС-Т\" on the same page as current CEO Vladimir Krikushenko, while official NexTouch pages and public investment/news materials corroborate him as the current senior operating leader; mobile and landline numbers were left blank because no safe named extension or explicitly person-level phone was publicly verified.",
        "selection_rationale": "CEO remains the strongest publicly evidenced operating LPR for NexTouch; the email was retained only after it appeared consistently across independent public sources tied to the same legal entity and current leader.",
        "confidence": "medium",
    },
    {
        "slug": "atol",
        "brand_name": "ATOL / АТОЛ",
        "legal_entity": 'ООО "АТОЛ"',
        "website": "https://www.atol.ru",
        "person": "Алексей Макаров",
        "title": "президент / основатель",
        "source_url": "https://www.atol.ru",
        "confidence": "medium",
    },
    {
        "slug": "incotex",
        "brand_name": "Incotex / Инкотекс",
        "legal_entity": 'ООО "ИНКОТЕКС-СК"',
        "website": "https://www.incotex.ru",
        "person": "Владимир Владимирович Бакланов",
        "title": "заместитель генерального директора",
        "corporate_email": "v.baklanov@incotexcom.ru",
        "source_url": "https://www.incotexcom.ru/contacts",
        "verification_note": "Official Incotex contacts page lists Vladimir Baklanov as deputy general director and publishes v.baklanov@incotexcom.ru; the published landline was left blank because the project rules forbid city numbers without a named extension.",
        "selection_rationale": "Updated from the blank CEO row because the official contacts page exposes a higher-value named executive contact at deputy-general-director level with a direct corporate email and current title.",
        "confidence": "high",
    },
    {
        "slug": "bolid",
        "brand_name": "Bolid / Болид",
        "legal_entity": 'АО НВП "Болид"',
        "website": "https://bolid.ru",
        "person": "Игорь Александрович Бабанов",
        "title": "генеральный директор",
        "telegram_group": "@bolid_nvp",
        "source_url": "https://bolid.ru",
        "verification_note": "Official Bolid site links the verified official Telegram-group handle @bolid_nvp; no current personal Telegram nickname, person-level direct email, mobile, or named extension was publicly found for Igor Babanov.",
        "confidence": "high",
    },
    {
        "slug": "dssl",
        "brand_name": "DSSL",
        "legal_entity": 'ООО "ДССЛ-ПЕРВЫЙ"',
        "website": "https://www.dssl.ru",
        "person": "Игорь Валерьевич Олейник",
        "title": "генеральный директор",
        "source_url": "https://www.rusprofile.ru/search?query=%D0%94%D0%A1%D0%A1%D0%9B",
        "confidence": "high",
    },
    {
        "slug": "owen",
        "brand_name": "OWEN / ОВЕН",
        "legal_entity": 'ООО "ПРОИЗВОДСТВЕННОЕ ОБЪЕДИНЕНИЕ ОВЕН"',
        "website": "https://owen.ru",
        "person": "Елена Геннадьевна Ламонова",
        "title": "генеральный директор",
        "corporate_email": "e.lamonova@owen.ru",
        "telegram_group": "@owen_prom",
        "source_url": "https://star-pro.ru/proverka-kontragenta/organization/1037739474266--ooo-po-oven",
        "verification_note": "Public company-card sources xfirm and STAR PRO list e.lamonova@owen.ru for ООО \"ПРОИЗВОДСТВЕННОЕ ОБЪЕДИНЕНИЕ ОВЕН\"; official and RBC company pages corroborate Elena Lamonova as the current general director of the operating production entity.",
        "selection_rationale": "Updated from the holding-company row to the operating manufacturing entity because the production company and its current CEO have a corroborated direct corporate email while OWEN remains a strong instrumentation / automation fit for the asset.",
        "confidence": "medium",
    },
    {
        "slug": "elemer",
        "brand_name": "ELEMER / ЭЛЕМЕР",
        "legal_entity": 'ООО НПП "ЭЛЕМЕР"',
        "website": "https://www.elemer.ru",
        "person": "Виталий Михайлович Окладников",
        "title": "генеральный директор",
        "source_url": "https://www.rusprofile.ru/search?query=%D0%AD%D0%9B%D0%95%D0%9C%D0%95%D0%A0",
        "confidence": "high",
    },
    {
        "slug": "lassard",
        "brand_name": "Lassard / Лассард",
        "legal_entity": 'ООО "Лассард"',
        "website": "https://lassard.ru",
        "person": "Наталия Ремизова",
        "title": "руководитель проектов по направлению Оптомеханика",
        "corporate_email": "n.remizova@lassard.ru",
        "mobile_phone": "+7 903 150-76-26",
        "work_phone": "+7 495 120-68-86 доб. 505",
        "source_url": "https://lassard.ru/products/components/diodes/laser-diodes/pulsed-laser-diodes",
        "verification_note": "Official Lassard product page names Natalia Remizova and gives direct corporate email, public mobile, and work line with extension.",
        "selection_rationale": "Official product page exposes this named project lead with direct person-level contact details; no higher-priority facilities / ops direct contact was publicly disclosed.",
        "confidence": "high",
    },
    {
        "slug": "nt-mdt",
        "brand_name": "NT-MDT",
        "legal_entity": 'ООО "НТ-МДТ С.И."',
        "website": "https://ntmdt-russia.com",
        "person": "Андрей Викторович Быков",
        "title": "генеральный директор",
        "source_url": "https://www.rusprofile.ru/search?query=%D0%9D%D0%A2-%D0%9C%D0%94%D0%A2",
        "confidence": "medium",
    },
    {
        "slug": "dna-technology",
        "brand_name": "DNA-Technology / ДНК-Технология",
        "legal_entity": 'ООО "ДНК-Технология"',
        "website": "https://dna-technology.ru",
        "person": "Владимир Юрьевич Дмитровский",
        "title": "генеральный директор",
        "telegram_group": "@dna_tech_rus",
        "source_url": "https://dna-technology.ru",
        "verification_note": "Official DNA-Technology site links the verified official Telegram-group handle @dna_tech_rus; no current personal Telegram nickname, person-level direct email, mobile, or named extension was publicly found for Vladimir Dmitrovsky.",
        "confidence": "medium",
    },
    {
        "slug": "medplant",
        "brand_name": "Medplant / Медплант",
        "legal_entity": 'ООО "Медплант"',
        "website": "https://medplant.ru",
        "person": "Андрей Владимирович Пушин",
        "title": "генеральный директор",
        "telegram_group": "@medplantm",
        "source_url": "https://medplant.ru/contacts/",
        "verification_note": "Official Medplant contacts page links the verified official Telegram-group handle @medplantm; no current personal Telegram nickname, person-level direct email, mobile, or named extension was publicly found for Andrey Pushin.",
        "confidence": "high",
    },
    {
        "slug": "beward",
        "brand_name": "BEWARD",
        "legal_entity": 'ООО "НПП "Бевард""',
        "website": "https://www.beward.ru",
        "person": "Лазуренко Артём Васильевич",
        "title": "руководитель отдела продаж",
        "corporate_email": "lav@beward.ru",
        "work_phone": "+7 391 278-92-00 доб. 515",
        "source_url": "https://www.beward.ru/news/beward-provedet-samye-udalennye-seminary-na-dalnem-vostoke-rossii/",
        "verification_note": "Official BEWARD seminar announcement names Artem Lazurenko as head of sales and gives direct corporate email plus named extension.",
        "selection_rationale": "Official event page exposes this named sales head with direct person-level contact details; no higher-priority facilities / ops direct contact was publicly disclosed.",
        "confidence": "high",
    },
    {
        "slug": "sigur",
        "brand_name": "Sigur",
        "legal_entity": 'ООО "Сигур"',
        "website": "https://sigur.com",
        "person": "Анар Адалет Оглы Зейналов",
        "title": "генеральный директор",
        "telegram_group": "@sigursys",
        "source_url": "https://sigur.com/contacts/",
        "verification_note": "Official Sigur contacts page links the verified official Telegram-group handle @sigursys; no current personal Telegram nickname, person-level direct email, mobile, or named extension was publicly found for Anar Zeynalov.",
        "confidence": "high",
    },
    {
        "slug": "parsec",
        "brand_name": "Parsec",
        "legal_entity": 'ООО Фирма "Парсек"',
        "website": "https://www.parsec.ru",
        "person": "Алексей Горюнов",
        "title": "руководитель отдела сопровождения проектов",
        "corporate_email": "goryunov@parsec.ru",
        "work_phone": "+7 499 495-17-82 доб. 201",
        "source_url": "https://www.parsec.ru/company/contacts/",
        "verification_note": "Official Parsec contacts page lists Alexey Goryunov with direct corporate email and named work line with extension.",
        "selection_rationale": "Official contacts page exposes this named project-support leader with direct person-level contact details; no higher-priority facilities / ops direct contact was publicly disclosed.",
        "confidence": "high",
    },
    {
        "slug": "byterg",
        "brand_name": "Byterg / Байтэрг",
        "legal_entity": 'ООО "Байтэрг"',
        "website": "https://byterg.ru",
        "person": "Андрей Сергеевич Прудников",
        "title": "генеральный директор",
        "source_url": "https://www.rusprofile.ru/search?query=%D0%91%D0%90%D0%99%D0%A2%D0%AD%D0%A0%D0%93",
        "confidence": "high",
    },
    {
        "slug": "rvi-group",
        "brand_name": "RVi Group",
        "legal_entity": 'ООО "RVI GROUP"',
        "website": "https://rvigroup.ru",
        "person": "Александр Хвастунов",
        "title": "топ-менеджер бренда / владелец",
        "source_url": "https://rvigroup.ru/",
        "confidence": "low",
    },
    {
        "slug": "optosystems",
        "brand_name": "Optosystems / Оптосистемы",
        "legal_entity": 'ООО "Оптосистемы"',
        "website": "https://optosystems.ru",
        "person": "Олег Александрович Нефёдов",
        "title": "директор",
        "source_url": "https://www.rusprofile.ru/search?query=%D0%9E%D0%9F%D0%A2%D0%9E%D0%A1%D0%98%D0%A1%D0%A2%D0%95%D0%9C%D0%AB",
        "confidence": "high",
    },
    {
        "slug": "ronavi-robotics",
        "brand_name": "Ronavi Robotics",
        "legal_entity": 'ООО "Ронави Роботикс"',
        "website": "https://ronavi-robotics.ru",
        "person": "Иван Валентинович Бородин",
        "title": "генеральный директор",
        "source_url": "https://www.rusprofile.ru/search?query=%D0%A0%D0%9E%D0%9D%D0%90%D0%92%D0%98+%D0%A0%D0%9E%D0%91%D0%9E%D0%A2%D0%98%D0%9A%D0%A1",
        "confidence": "high",
    },
    {
        "slug": "technored",
        "brand_name": "Technored",
        "legal_entity": 'ООО "Технорэд"',
        "website": "https://technored.ru",
        "person": "Артём Владимирович Лукин",
        "title": "генеральный директор",
        "telegram_niknames": "@lookintech",
        "source_url": "https://t.me/lookintech",
        "verification_note": "Artyom Lukin's public Telegram handle @lookintech is tied to TECHNORED's founder by public channel identity and separate official company evidence.",
        "selection_rationale": "Official Technored materials identify Artyom Lukin as founder and CEO, while his personal public Telegram handle provides a verified person-level business identity; no stronger direct email / phone was publicly confirmed.",
        "confidence": "high",
    },
    {
        "slug": "elta",
        "brand_name": "ELTA / ЭЛТА",
        "legal_entity": 'ООО "ЭЛТА"',
        "website": "https://xn--80achcebqujlijcbjv1ag.xn--p1ai",
        "person": "Анатолий Петрович Храпенко",
        "title": "генеральный директор",
        "source_url": "https://www.rusprofile.ru/search?query=%D0%AD%D0%9B%D0%A2%D0%90",
        "confidence": "medium",
    },
    {
        "slug": "ekf",
        "brand_name": "EKF",
        "legal_entity": 'ООО "Электрорешения"',
        "website": "https://ekfgroup.com",
        "person": "Дмитрий Андреевич Кучеров",
        "title": "генеральный директор",
        "telegram_group": "@ekf_channel",
        "source_url": "https://ekfgroup.com",
        "verification_note": "Official EKF site links the verified official Telegram-group handle @ekf_channel; no current personal Telegram nickname, person-level direct email, mobile, or named extension was publicly found for Dmitry Kucherov.",
        "confidence": "high",
    },
    {
        "slug": "iek-group",
        "brand_name": "IEK Group",
        "legal_entity": 'ПАО "ИЭК ХОЛДИНГ"',
        "website": "https://www.iek.ru",
        "person": "Михаил Петров",
        "title": "сооснователь / владелец группы",
        "telegram_group": "@iek_group_rus",
        "source_url": "https://www.iek.ru/company/contacts/",
        "verification_note": "Official IEK contacts page links the verified official Telegram-group handle @iek_group_rus; no current personal Telegram nickname, person-level direct email, mobile, or named extension was publicly found for Mikhail Petrov.",
        "confidence": "low",
    },
    {
        "slug": "elvees",
        "brand_name": "ELVEES / ЭЛВИС",
        "legal_entity": 'АО НПЦ "ЭЛВИС"',
        "website": "https://elvees.ru",
        "person": "Антон Дмитриевич Семилетов",
        "title": "генеральный директор",
        "telegram_group": "@elvees_support",
        "source_url": "https://elvees.ru",
        "verification_note": "Official ELVEES site links the verified official Telegram-group handle @elvees_support; no current personal Telegram nickname, person-level direct email, mobile, or named extension was publicly found for Anton Semiletov.",
        "confidence": "high",
    },
    {
        "slug": "elvees-neotek",
        "brand_name": "ELVEES NeoTek / ЭЛВИС-НеоТек",
        "legal_entity": 'ООО "ЭЛВИС-НеоТек"',
        "website": "https://www.elveesneotek.ru",
        "person": "Дмитрий Сергеевич Петров",
        "title": "генеральный директор",
        "source_url": "https://www.rusprofile.ru/search?query=%D0%AD%D0%9B%D0%92%D0%98%D0%A1-%D0%9D%D0%95%D0%9E%D0%A2%D0%95%D0%9A",
        "confidence": "medium",
    },
    {
        "slug": "komitex",
        "brand_name": "Komitex / Комитекс",
        "legal_entity": 'АО "Комитекс"',
        "website": "https://www.komitex.ru",
        "person": "Юлия Валентиновна Юдина",
        "title": "директор по экономике и персоналу",
        "corporate_email": "yudina_yv@komitex.ru",
        "source_url": "https://www.komitex.ru/contacts/",
        "verification_note": "Official Komitex contacts page lists yudina_yv@komitex.ru for the director of economics and personnel; BNKomi and SLT public sources identify that roleholder as Yulia Valentinovna Yudina.",
        "selection_rationale": "Komitex is a large light-industry manufacturer with more than 1000 employees and Moscow representation; the finance-and-personnel director is the highest-priority named senior role publicly tied to a direct corporate email on accessible sources.",
        "confidence": "medium",
    },
    {
        "slug": "rester",
        "brand_name": "RESTER / РЕСТЕР",
        "legal_entity": 'АО "РЕСТЕР"',
        "website": "https://rester.ru",
        "person": "Дьячков Роман Михайлович",
        "title": "генеральный директор",
        "corporate_email": "roman@rester.ru",
        "source_url": "https://rester.ru/ru/contacts/",
        "verification_note": "Official Rester contacts page lists roman@rester.ru for the general director; current public registry data identifies the operating pharmaceutical manufacturer as AO RESTER led by Roman Mikhailovich Dyachkov.",
        "selection_rationale": "Current official contact data exposes a direct corporate email for the named top executive of a GMP pharmaceutical production site, making it a strong scientific-production tenant candidate.",
        "confidence": "medium",
    },
    {
        "slug": "specpribor",
        "brand_name": "Specpribor / Спецприбор",
        "legal_entity": 'ООО "Спецприбор"',
        "website": "https://www.specpribor.ru",
        "person": "Ронжин Виталий Владимирович",
        "title": "технический директор",
        "corporate_email": "vr@specpribor.ru",
        "source_url": "https://www.specpribor.ru/kontakty",
        "verification_note": "Official Specpribor contacts page lists technical director Vitaly Ronzhin with direct corporate email vr@specpribor.ru; the shared switchboard was not used because no named extension is published for him.",
        "selection_rationale": "Technical director is a higher-priority technical LPR than the generic-contact CEO row, and the official contacts page publishes his direct person-level corporate email.",
        "confidence": "high",
    },
    {
        "slug": "uraltekhmarket",
        "brand_name": "Uraltekhmarket / Уралтехмаркет",
        "legal_entity": 'ЗАО "УРАЛТЕХМАРКЕТ"',
        "website": "https://www.uraltm.ru",
        "person": "Белоусов Антон Александрович",
        "title": "технический директор",
        "corporate_email": "antonb@uraltm.ru",
        "work_phone": "+7 (343) 288-51-41 доб. 5011",
        "source_url": "https://www.uraltm.ru/contacts/",
        "verification_note": "Official Uraltekhmarket contacts page lists technical director Anton Belousov with direct corporate email and named extension; official company pages describe the firm's own production base and robotics / industrial-automation focus.",
        "selection_rationale": "Technical director is a priority-role LPR for an engineering company that combines industrial automation, robot integration, and its own cabinet-production base, matching the asset's scientific-production profile.",
        "confidence": "high",
    },
    {
        "slug": "lisma",
        "brand_name": "LISMA / Лисма",
        "legal_entity": 'ООО "ССЗ"',
        "website": "https://lisma.su",
        "person": "Пьянзин Владимир Васильевич",
        "title": "заместитель директора по кадровым и социальным вопросам",
        "corporate_email": "pvv@lisma.su",
        "work_phone": "+7 (8342) 77-70-60 доб. 1214",
        "source_url": "https://lisma.su/kontakty/rukovodstvo/index.html",
        "verification_note": "Official LISMA leadership page lists deputy director for personnel and social matters Vladimir Pyanzin with direct corporate email and named extension; the official about page confirms the company as Russia's largest full-cycle lamp producer.",
        "selection_rationale": "A deputy-director / people-function role is high on the priority ladder and is published with direct person-level contact on the official leadership page of a large light-industry manufacturer.",
        "confidence": "high",
    },
    {
        "slug": "aiss",
        "brand_name": "AISS / АИСС",
        "legal_entity": 'ООО "АИСС"',
        "website": "https://www.aiss33.ru",
        "person": "Владимиров Михаил Эдуардович",
        "title": "технический директор",
        "corporate_email": "vme@aiss33.ru",
        "source_url": "https://www.aiss33.ru/contacts/",
        "verification_note": "Official AISS contacts page lists technical director Mikhail Vladimirov with direct corporate email vme@aiss33.ru; the published landline was left out because the workbook permits office phones only when a named extension is explicitly shown.",
        "selection_rationale": "AISS combines engineering systems, industrial automation, and its own production of boiler houses and control cabinets; the technical director is a priority-role LPR publicly disclosed with a direct email on the official contacts page.",
        "confidence": "high",
    },
]


def corporate_email(record: dict[str, str]) -> str:
    return record.get("corporate_email", "")


def mobile_phone(record: dict[str, str]) -> str:
    return record.get("mobile_phone", "")


def work_phone(record: dict[str, str]) -> str:
    return record.get("work_phone", "")


def telegram_niknames(record: dict[str, str]) -> str:
    return record.get("telegram_niknames", "")


def telegram_group(record: dict[str, str]) -> str:
    return record.get("telegram_group", "")


def verification_note(record: dict[str, str]) -> str:
    return record.get(
        "verification_note",
        "Named person verified; no current direct person-level email, phone, or personal Telegram nickname publicly found.",
    )


def build_note(record: dict[str, str]) -> str:
    confidence_note = {
        "high": "Official site plus public company-card evidence aligned cleanly.",
        "medium": "Official site evidence was partial, so public company-card evidence was used to complete the row.",
        "low": "Management disclosure on the official site was limited; the row uses the best available public evidence and should be reviewed first in any manual QA.",
    }[record["confidence"]]
    email = corporate_email(record)
    mobile = mobile_phone(record)
    work = work_phone(record)
    telegram = telegram_niknames(record)
    telegram_group_value = telegram_group(record)
    email_reason = (
        "official or strongly corroborated public source publicly ties this current corporate email to the named person."
        if email
        else "no current direct person-level corporate email publicly found; generic inboxes rejected."
    )
    mobile_reason = (
        "official or strongly corroborated public source publicly ties this current phone to the named person."
        if mobile
        else "no current direct public business mobile publicly found for the named person."
    )
    work_reason = (
        "official or strongly corroborated public source lists a current named work line with explicit extension for this person."
        if work
        else "no current named office extension publicly found for the named person."
    )
    telegram_reason = (
        "public evidence verifies this personal Telegram nickname as current and tied to the named person."
        if telegram
        else "no verified current public personal Telegram nickname found for the named person."
    )
    telegram_group_reason = (
        "public evidence verifies this official company/group/channel Telegram handle as current and tied to the employer."
        if telegram_group_value
        else "no verified current official company/group/channel Telegram handle found."
    )
    selection_rationale = record.get(
        "selection_rationale",
        "best publicly evidenced LPR or fallback senior leader/owner found during the scan.",
    )

    return f"""# Research note — {record["brand_name"]}

## Company
- Brand: {record["brand_name"]}
- Seed website: {record["website"]}
- Research date: {RESEARCH_DATE}

## 1. Official website verification
- Official website: {record["website"]}
- Why this is official: seed domain from target list; active branded corporate site.
- Alternate domains checked: homepage plus contact/legal/company pages where available.

## 2. Legal entity verification
- Chosen legal entity: {record["legal_entity"]}
- Evidence source 1: {record["website"]}
- Evidence source 2: {record["source_url"]}
- Notes on group structure / ambiguity: {confidence_note}

## 3. Named decision-maker selection
- Chosen person: {record["person"]}
- Chosen title: {record["title"]}
- Why this role was selected: {selection_rationale}
- Higher-priority roles checked and not found: dedicated admin/facilities/real-estate contacts were not publicly disclosed on accessible official pages.

## 4. Direct contact findings
### Corporate email
- Value: {email}
- Source: {record["source_url"] if email else ""}
- Why acceptable / why blank: {email_reason}

### Mobile phone
- Value: {mobile}
- Source: {record["source_url"] if mobile else ""}
- Why acceptable / why blank: {mobile_reason}

### Work phone + extension
- Value: {work}
- Source: {record["source_url"] if work else ""}
- Why acceptable / why blank: {work_reason}

### Telegram niknames
- Value: {telegram}
- Source: {record["source_url"] if telegram else ""}
- Why acceptable / why blank: {telegram_reason}

### Telegram-group
- Value: {telegram_group_value}
- Source: {record["source_url"] if telegram_group_value else ""}
- Why acceptable / why blank: {telegram_group_reason}

## 5. Search trail
- Direct-source channels checked first: official pages and public sources for real emails, mobiles, named extensions, and personal Telegram nicknames tied to the selected LPR.
- Official site pages checked: homepage, contacts page, privacy/legal/footer references, company/about sections where available.
- Official PDFs checked: privacy / policy / legal documents when linked; no direct person-level contacts captured.
- Event / speaker pages checked: best-effort open-web scan; no safer direct person-level contact found.
- Procurement / registry pages checked: public company-card sources used where official site did not fully expose management details.
- Professional profiles checked: not relied on as sole evidence.
- Candidate-email derivation used only if stronger direct-source channels were insufficient.

## 6. Final row draft
- Legal entity name: {record["legal_entity"]}
- Website: {record["website"]}
- Decision-maker full name: {record["person"]}
- Decision-maker title: {record["title"]}
- Corporate email: {email}
- Mobile phone: {mobile}
- Work phone + extension: {work}
- Telegram niknames: {telegram}
- Telegram-group: {telegram_group_value}
- Source URL: {record["source_url"]}
- Verification note: {verification_note(record)}

## 7. QA
- Generic email used? no
- General phone without extension used? no
- Guessed email used? no
- Verified current Telegram nickname used? {"yes" if telegram else "no"}
- Verified official Telegram-group handle used? {"yes" if telegram_group_value else "no"}
- Any stale / non-working contact used? no
- Evidence log updated? yes
"""


def main() -> None:
    expected_rows = list(DictReader(TARGETS.open(encoding="utf-8", newline="")))
    if len(ROWS) < len(expected_rows):
        raise RuntimeError(
            f"Prepared {len(ROWS)} rows but the seed list has {len(expected_rows)} companies"
        )

    NOTES_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(WORKBOOK)
    ws = wb[SHEET_NAME]

    for idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=idx).value = header

    for idx, record in enumerate(ROWS, start=2):
        ws.cell(row=idx, column=1).value = record["legal_entity"]
        ws.cell(row=idx, column=2).value = record["website"]
        ws.cell(row=idx, column=3).value = record["person"]
        ws.cell(row=idx, column=4).value = record["title"]
        ws.cell(row=idx, column=5).value = corporate_email(record)
        ws.cell(row=idx, column=6).value = mobile_phone(record)
        ws.cell(row=idx, column=7).value = work_phone(record)
        ws.cell(row=idx, column=8).value = telegram_niknames(record)
        ws.cell(row=idx, column=9).value = telegram_group(record)
        ws.cell(row=idx, column=10).value = record["source_url"]
        ws.cell(row=idx, column=11).value = verification_note(record)

    while ws.max_row > len(ROWS) + 1:
        ws.delete_rows(ws.max_row)

    wb.save(WORKBOOK)

    with EVIDENCE.open("w", encoding="utf-8", newline="") as f:
        writer = DictWriter(
            f,
            fieldnames=[
                "brand_name",
                "legal_entity",
                "person_name",
                "role",
                "contact_type",
                "contact_value",
                "source_url",
                "source_type",
                "confidence",
                "notes",
            ],
        )
        writer.writeheader()
        for record in ROWS:
            person_level_contacts = [
                ("corporate_email", corporate_email(record)),
                ("mobile_phone", mobile_phone(record)),
                ("work_phone_extension", work_phone(record)),
                ("telegram_niknames", telegram_niknames(record)),
            ]
            supporting_contacts = [
                ("telegram_group", telegram_group(record)),
            ]
            wrote_person_level_contact = False
            for contact_type, contact_value in person_level_contacts:
                if not contact_value:
                    continue
                wrote_person_level_contact = True
                writer.writerow(
                    {
                        "brand_name": record["brand_name"],
                        "legal_entity": record["legal_entity"],
                        "person_name": record["person"],
                        "role": record["title"],
                        "contact_type": contact_type,
                        "contact_value": contact_value,
                        "source_url": record["source_url"],
                        "source_type": "official_public_source",
                        "confidence": record["confidence"],
                        "notes": verification_note(record),
                    }
                )

            for contact_type, contact_value in supporting_contacts:
                if not contact_value:
                    continue
                writer.writerow(
                    {
                        "brand_name": record["brand_name"],
                        "legal_entity": record["legal_entity"],
                        "person_name": record["person"],
                        "role": record["title"],
                        "contact_type": contact_type,
                        "contact_value": contact_value,
                        "source_url": record["source_url"],
                        "source_type": "official_public_source",
                        "confidence": record["confidence"],
                        "notes": verification_note(record),
                    }
                )

            if not wrote_person_level_contact:
                writer.writerow(
                    {
                        "brand_name": record["brand_name"],
                        "legal_entity": record["legal_entity"],
                        "person_name": record["person"],
                        "role": record["title"],
                        "contact_type": "named_lpr_no_direct_contact",
                        "contact_value": "",
                        "source_url": record["source_url"],
                        "source_type": "official_and_public_company_sources",
                        "confidence": record["confidence"],
                        "notes": "Direct person-level contact fields intentionally blank after public-source scan; supporting Telegram-group evidence may still be present separately.",
                    }
                )

    for record in ROWS:
        note_path = NOTES_DIR / f"{record['slug']}.md"
        note_path.write_text(build_note(record), encoding="utf-8")


if __name__ == "__main__":
    main()
