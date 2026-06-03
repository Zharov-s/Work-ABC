from __future__ import annotations

from openpyxl import load_workbook
from urllib.parse import urlparse
from pathlib import Path
import subprocess
import re
import json

WORKBOOK = Path("output/mitino_target_companies_lpr_direct_contacts.xlsx")
OUT_JSON = Path("output/found_new_lpr_emails.json")

GENERIC = {
    "info",
    "sales",
    "office",
    "support",
    "mail",
    "contact",
    "zakaz",
    "commerce",
    "pr",
    "press",
    "media",
    "marketing",
    "hr",
    "career",
    "corp",
    "admin",
    "reception",
    "hotline",
    "personal",
    "service",
    "servis",
    "secretary",
    "sekretar",
    "operator",
    "shop",
    "sale",
    "snab",
    "buh",
    "pto",
    "nocorruption",
}

FREE = {
    "mail.ru",
    "gmail.com",
    "yandex.ru",
    "yandex.com",
    "bk.ru",
    "inbox.ru",
    "list.ru",
    "hotmail.com",
    "outlook.com",
}

PATHS = [
    "",
    "/contacts",
    "/contact",
    "/kontakty",
    "/contacts/",
    "/contact/",
    "/kontakty/",
    "/about",
    "/about/",
    "/company",
    "/company/",
    "/about-us",
    "/about-us/",
    "/o-kompanii",
    "/o-kompanii/",
    "/team",
    "/team/",
    "/management",
    "/management/",
    "/leadership",
    "/leadership/",
    "/rukovodstvo",
    "/rukovodstvo/",
]

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")


def load_rows() -> tuple[list[tuple[str, str, str]], set[str]]:
    wb = load_workbook(WORKBOOK, read_only=True)
    ws = wb["Контрагенты"]
    rows: list[tuple[str, str, str]] = []
    existing: set[str] = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        for e in str(r[14] or "").split("\n"):
            e = e.strip().lower()
            if e:
                existing.add(e)
        company = str(r[0] or "").strip()
        person = str(r[10] or "").strip()
        site = str(r[15] or "").strip()
        if company and person and site:
            rows.append((company, person, site))
    return rows, existing


def fetch(url: str) -> str:
    try:
        out = subprocess.run(
            ["curl", "-L", url],
            capture_output=True,
            text=True,
            timeout=12,
        )
    except Exception:
        return ""
    if out.returncode != 0:
        return ""
    return out.stdout or ""


def normalize_email(email: str) -> str:
    return email.lower().strip(".,;:()[]<>\"'")


def is_allowed(email: str, existing: set[str], seen: set[str]) -> bool:
    eml = normalize_email(email)
    if not eml or eml in existing or eml in seen or "@" not in eml:
        return False
    local, domain = eml.split("@", 1)
    if domain in FREE:
        return False
    if local in GENERIC:
        return False
    if any(tok in local for tok in ("noreply", "no-reply", "do-not-reply")):
        return False
    return True


def main() -> None:
    rows, existing = load_rows()
    found: list[dict[str, str]] = []
    seen: set[str] = set()

    for idx, (company, person, site) in enumerate(rows, start=1):
        if not re.match(r"^https?://", site, re.I):
            site = "http://" + site
        parsed = urlparse(site)
        base = f"{parsed.scheme}://{parsed.netloc}"
        surname = person.split()[0].lower()
        person_l = person.lower()

        hits: list[tuple[str, str]] = []
        for path in PATHS:
            url = base + path
            text = fetch(url)
            if not text:
                continue
            lower = text.lower()
            if surname not in lower and person_l not in lower:
                continue
            for em in set(EMAIL_RE.findall(text)):
                if is_allowed(em, existing, seen):
                    hits.append((normalize_email(em), url))
            if hits:
                break

        for eml, url in hits:
            found.append(
                {
                    "company": company,
                    "person": person,
                    "email": eml,
                    "url": url,
                }
            )
            seen.add(eml)

        if idx % 100 == 0:
            print(f"scanned {idx} found {len(found)}", flush=True)

    OUT_JSON.write_text(json.dumps(found, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"total_found {len(found)}")
    for item in found[:200]:
        print(f"{item['email']} | {item['person']} | {item['company']} | {item['url']}")


if __name__ == "__main__":
    main()
