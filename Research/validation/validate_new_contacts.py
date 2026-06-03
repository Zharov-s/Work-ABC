from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Sequence
from urllib.parse import urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "output" / "new_contacts.xlsx"
DEFAULT_MEMORY = ROOT / "output" / "found_contacts_memory.jsonl"
DEFAULT_EVIDENCE = ROOT / "output" / "evidence_log.csv"
BLOCKED_EMAILS = ROOT / "validation" / "blocked_generic_emails.txt"
SUSPICIOUS_EMAILS = ROOT / "validation" / "suspicious_email_patterns.txt"
SUSPICIOUS_PHONES = ROOT / "validation" / "suspicious_phone_patterns.txt"

EXPECTED_SHEET = "new_contacts"
EXPECTED_HEADERS = [
    "Наименование",
    "Сайт",
    "ФИО",
    "Должность",
    "Емайл",
    "Мобильный телефон или городской с добавочным",
]
PERSONAL_EMAIL_DOMAINS = {
    "gmail.com",
    "googlemail.com",
    "mail.ru",
    "inbox.ru",
    "list.ru",
    "bk.ru",
    "internet.ru",
    "yandex.ru",
    "yandex.com",
    "ya.ru",
    "icloud.com",
    "me.com",
    "mac.com",
    "hotmail.com",
    "outlook.com",
    "live.com",
    "msn.com",
    "yahoo.com",
    "protonmail.com",
    "rambler.ru",
}
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
EXTENSION_RE = re.compile(r"(доб\.?|добавочн|ext\.?|extension|вн\.?)\s*[:#№-]?\s*\d+", re.IGNORECASE)


def normalize(value: object) -> str:
    return (value or "").strip() if isinstance(value, str) else str(value or "").strip()


def normalize_company(value: object) -> str:
    return re.sub(r"\s+", " ", normalize(value)).lower()


def normalize_website(value: object) -> str:
    raw = normalize(value).lower().strip()
    if not raw:
        return ""
    candidate = raw if "://" in raw else f"https://{raw}"
    parsed = urlparse(candidate)
    host = (parsed.netloc or parsed.path).lower().strip().strip("/")
    if host.startswith("www."):
        host = host[4:]
    path = parsed.path.strip("/") if parsed.netloc else ""
    return f"{host}/{path}" if path else host


def load_list(path: Path) -> list[str]:
    if not path.exists():
        return []
    return [line.strip().lower() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def load_patterns(path: Path) -> list[re.Pattern[str]]:
    patterns: list[re.Pattern[str]] = []
    if not path.exists():
        return patterns
    for line in path.read_text(encoding="utf-8").splitlines():
        value = line.strip()
        if not value:
            continue
        patterns.append(re.compile(value, re.IGNORECASE))
    return patterns


def split_contacts(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[;,\n]", value) if item.strip()]


def split_phone_channels(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[;\n]", value) if item.strip()]


def digit_count(value: str) -> int:
    return len(re.sub(r"\D", "", value))


def has_extension_marker(value: str) -> bool:
    return bool(EXTENSION_RE.search(value))


def is_probable_russian_mobile(value: str) -> bool:
    digits = re.sub(r"\D", "", value)
    if len(digits) == 10 and digits.startswith("9"):
        return True
    if len(digits) == 11 and (digits.startswith("79") or digits.startswith("89")):
        return True
    return False


def email_domain(value: str) -> str:
    return value.lower().rsplit("@", 1)[1] if "@" in value else ""


def website_domain(value: str) -> str:
    return normalize_website(value).split("/", 1)[0]


def domains_look_related(email_host: str, site_host: str) -> bool:
    if not email_host or not site_host:
        return True
    return (
        email_host == site_host
        or email_host.endswith(f".{site_host}")
        or site_host.endswith(f".{email_host}")
    )


def is_generic_email(value: str, blocked: list[str], patterns: list[re.Pattern[str]]) -> bool:
    lowered = value.lower()
    if any(token in lowered for token in blocked):
        return True
    return any(pattern.search(lowered) for pattern in patterns)


def has_suspicious_phone(value: str, patterns: list[re.Pattern[str]]) -> bool:
    return any(pattern.search(value) for pattern in patterns)


def load_memory_index(path: Path) -> tuple[set[str], set[str]]:
    companies: set[str] = set()
    websites: set[str] = set()
    if not path.exists():
        return companies, websites

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        try:
            record = json.loads(line)
        except json.JSONDecodeError:
            continue
        companies.add(normalize_company(record.get("company_name", "")))
        websites.add(normalize_website(record.get("website", "")))
    return companies, websites


def load_evidence_index(path: Path) -> tuple[set[str], set[tuple[str, str]]]:
    companies: set[str] = set()
    company_people: set[tuple[str, str]] = set()
    if not path.exists():
        return companies, company_people

    with path.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            company = normalize_company(row.get("legal_entity", ""))
            person = normalize(row.get("person_name", "")).lower()
            if company:
                companies.add(company)
            if company and person:
                company_people.add((company, person))
    return companies, company_people


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate accelerated output/new_contacts.xlsx before preparation or mailing."
    )
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--memory", default=str(DEFAULT_MEMORY))
    parser.add_argument("--evidence", default=str(DEFAULT_EVIDENCE))
    parser.add_argument("--expected-rows", type=int, default=25)
    parser.add_argument(
        "--skip-memory-check",
        action="store_true",
        help="Skip duplicate-vs-memory enforcement. Useful only for local testing or resend scenarios.",
    )
    parser.add_argument(
        "--skip-evidence-check",
        action="store_true",
        help="Skip evidence-log coverage checks. Useful only for local testing.",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    workbook_path = Path(args.workbook).resolve()
    memory_path = Path(args.memory).resolve()
    evidence_path = Path(args.evidence).resolve()

    errors: list[str] = []
    warnings: list[str] = []

    blocked_emails = load_list(BLOCKED_EMAILS)
    suspicious_email_patterns = load_patterns(SUSPICIOUS_EMAILS)
    suspicious_phone_patterns = load_patterns(SUSPICIOUS_PHONES)

    memory_companies, memory_websites = set(), set()
    if not args.skip_memory_check:
        memory_companies, memory_websites = load_memory_index(memory_path)

    evidence_companies, evidence_company_people = set(), set()
    if not args.skip_evidence_check:
        evidence_companies, evidence_company_people = load_evidence_index(evidence_path)
        if not evidence_path.exists():
            errors.append(f"Evidence log not found: {evidence_path}")

    if not workbook_path.exists():
        print(f"ERROR: workbook not found: {workbook_path}")
        return 1

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if EXPECTED_SHEET not in wb.sheetnames:
        errors.append(f"Missing sheet: {EXPECTED_SHEET}")
    else:
        ws = wb[EXPECTED_SHEET]
        headers = [normalize(cell.value) for cell in ws[1][: len(EXPECTED_HEADERS)]]
        if headers != EXPECTED_HEADERS:
            errors.append(f"Header mismatch. Found: {headers}")

        seen_companies: set[str] = set()
        seen_websites: set[str] = set()
        seen_emails: set[str] = set()
        non_empty_rows = 0

        for idx, row in enumerate(
            ws.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS), values_only=True),
            start=2,
        ):
            values = [normalize(value) for value in row]
            if not any(values):
                continue
            non_empty_rows += 1
            (
                company_name,
                website,
                person_name,
                title,
                email,
                phone,
            ) = values

            if not company_name:
                errors.append(f"Row {idx}: missing `Наименование`")
            if not website:
                errors.append(f"Row {idx}: missing `Сайт`")
            elif "." not in website_domain(website):
                errors.append(f"Row {idx}: `Сайт` does not look like a valid official domain `{website}`")
            if not person_name:
                errors.append(f"Row {idx}: missing `ФИО`")
            if not title:
                errors.append(f"Row {idx}: missing `Должность`")
            if not email:
                errors.append(f"Row {idx}: missing `Емайл`")
            if not phone:
                errors.append(
                    f"Row {idx}: missing `Мобильный телефон или городской с добавочным`"
                )

            company_key = normalize_company(company_name)
            website_key = normalize_website(website)
            person_key = normalize(person_name).lower()

            if company_key in seen_companies:
                errors.append(f"Row {idx}: duplicate company `{company_name}` inside workbook")
            elif company_key:
                seen_companies.add(company_key)

            if website_key in seen_websites:
                errors.append(f"Row {idx}: duplicate website `{website}` inside workbook")
            elif website_key:
                seen_websites.add(website_key)

            email_tokens = split_contacts(email)
            if len(email_tokens) != 1:
                errors.append(
                    f"Row {idx}: `Емайл` must contain exactly one direct person-level address"
                )
            else:
                email_value = email_tokens[0]
                email_key = email_value.lower()
                if not EMAIL_RE.fullmatch(email_value):
                    errors.append(f"Row {idx}: invalid email format `{email_value}`")
                elif is_generic_email(email_value, blocked_emails, suspicious_email_patterns):
                    errors.append(f"Row {idx}: forbidden or suspicious email `{email_value}`")
                else:
                    domain = email_key.rsplit("@", 1)[1]
                    if domain in PERSONAL_EMAIL_DOMAINS:
                        errors.append(
                            f"Row {idx}: email `{email_value}` uses a personal mailbox domain, not a corporate one"
                        )
                if email_key in seen_emails:
                    errors.append(f"Row {idx}: duplicate email `{email_value}` inside workbook")
                else:
                    seen_emails.add(email_key)

            if phone and has_suspicious_phone(phone, suspicious_phone_patterns):
                errors.append(f"Row {idx}: suspicious phone value `{phone}`")

            phone_channels = split_phone_channels(phone)
            if phone and len(phone_channels) != 1:
                errors.append(
                    f"Row {idx}: phone field must contain exactly one verified channel, not multiple values"
                )
            elif phone:
                phone_value = phone_channels[0]
                if digit_count(phone_value) < 10:
                    errors.append(f"Row {idx}: phone value has fewer than 10 digits `{phone_value}`")
                if not is_probable_russian_mobile(phone_value) and not has_extension_marker(phone_value):
                    errors.append(
                        f"Row {idx}: phone must be a direct mobile or a named landline extension `{phone_value}`"
                    )

            if email:
                email_host = email_domain(email)
                site_host = website_domain(website)
                if email_host and site_host and not domains_look_related(email_host, site_host):
                    warnings.append(
                        f"Row {idx}: email domain `{email_host}` differs from website domain `{site_host}`; verify group-domain evidence"
                    )

            if not args.skip_memory_check:
                if company_key and company_key in memory_companies:
                    errors.append(
                        f"Row {idx}: company `{company_name}` already exists in `found_contacts_memory.jsonl`"
                    )
                if website_key and website_key in memory_websites:
                    errors.append(
                        f"Row {idx}: website `{website}` already exists in `found_contacts_memory.jsonl`"
                    )

            if not args.skip_evidence_check:
                if company_key and company_key not in evidence_companies:
                    errors.append(
                        f"Row {idx}: company `{company_name}` has no matching evidence-log entry"
                    )
                elif company_key and person_key and (company_key, person_key) not in evidence_company_people:
                    warnings.append(
                        f"Row {idx}: evidence log has the company but no direct person match for `{person_name}`"
                    )

        if non_empty_rows != args.expected_rows:
            errors.append(
                f"Workbook must contain exactly {args.expected_rows} filled rows; found {non_empty_rows}"
            )

    if errors:
        print("VALIDATION FAILED")
        print("=" * 80)
        for error in errors:
            print("ERROR:", error)
        if warnings:
            print("-" * 80)
            for warning in warnings:
                print("WARN:", warning)
        return 1

    print("VALIDATION PASSED")
    if warnings:
        print("=" * 80)
        for warning in warnings:
            print("WARN:", warning)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
