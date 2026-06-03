from __future__ import annotations

from pathlib import Path
import csv
import re

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "output" / "mitino_target_companies_lpr_direct_contacts.xlsx"
TARGETS = ROOT / "input" / "target_companies.csv"
EVIDENCE = ROOT / "output" / "evidence_log.csv"
BLOCKED_EMAILS = ROOT / "validation" / "blocked_generic_emails.txt"
SUSPICIOUS_EMAILS = ROOT / "validation" / "suspicious_email_patterns.txt"
SUSPICIOUS_PHONES = ROOT / "validation" / "suspicious_phone_patterns.txt"

EXPECTED_SHEET = "Direct LPR Contacts"
EXPECTED_HEADERS = [
    "Legal entity name",
    "Website",
    "Decision-maker full name",
    "Decision-maker title",
    "Corporate email",
    "Mobile phone",
    "Work phone + extension",
    "Telegram niknames",
    "Telegram-group",
    "Source URL",
    "Verification note",
]


def normalize(value) -> str:
    return (value or "").strip() if isinstance(value, str) else str(value or "").strip()


def load_list(path: Path) -> list[str]:
    if not path.exists():
        return []
    return [line.strip().lower() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def load_patterns(path: Path) -> list[re.Pattern[str]]:
    patterns: list[re.Pattern[str]] = []
    if not path.exists():
        return patterns
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        patterns.append(re.compile(line, re.IGNORECASE))
    return patterns


def has_suspicious_phone(value: str, patterns: list[re.Pattern[str]]) -> bool:
    return any(p.search(value) for p in patterns)


def is_generic_email(value: str, blocked: list[str], patterns: list[re.Pattern[str]]) -> bool:
    lowered = value.lower()
    if any(token in lowered for token in blocked):
        return True
    return any(p.search(lowered) for p in patterns)


def main() -> int:
    errors: list[str] = []
    warnings: list[str] = []

    blocked_emails = load_list(BLOCKED_EMAILS)
    suspicious_email_patterns = load_patterns(SUSPICIOUS_EMAILS)
    suspicious_phone_patterns = load_patterns(SUSPICIOUS_PHONES)

    if not WORKBOOK.exists():
        print(f"ERROR: workbook not found: {WORKBOOK}")
        return 1

    wb = load_workbook(WORKBOOK)
    if EXPECTED_SHEET not in wb.sheetnames:
        errors.append(f"Missing sheet: {EXPECTED_SHEET}")
    else:
        ws = wb[EXPECTED_SHEET]
        headers = [normalize(cell.value) for cell in ws[1][: len(EXPECTED_HEADERS)]]
        if headers != EXPECTED_HEADERS:
            errors.append(f"Header mismatch. Found: {headers}")

        seen_company_keys: set[tuple[str, str]] = set()
        contact_qualified = 0

        for idx, row in enumerate(
            ws.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS), values_only=True),
            start=2,
        ):
            vals = [normalize(x) for x in row]
            if len(vals) < len(EXPECTED_HEADERS):
                vals.extend([""] * (len(EXPECTED_HEADERS) - len(vals)))

            if not any(vals):
                continue

            (
                legal_entity,
                website,
                person_name,
                title,
                corporate_email,
                mobile_phone,
                work_phone_ext,
                telegram_nicknames,
                telegram_group,
                source_url,
                verification_note,
            ) = vals

            if not legal_entity:
                errors.append(f"Row {idx}: missing Legal entity name")
            if not website:
                warnings.append(f"Row {idx}: missing Website")
            if not person_name:
                errors.append(f"Row {idx}: missing Decision-maker full name")
            if not title:
                warnings.append(f"Row {idx}: missing Decision-maker title")
            if not source_url:
                errors.append(f"Row {idx}: missing Source URL")
            if not verification_note:
                errors.append(f"Row {idx}: missing Verification note")

            key = (legal_entity.lower(), website.lower())
            if key in seen_company_keys:
                errors.append(f"Row {idx}: duplicate company/legal-entity + website combination")
            else:
                seen_company_keys.add(key)

            if corporate_email:
                for email in [x.strip() for x in corporate_email.split(";") if x.strip()] + [
                    x.strip() for x in corporate_email.split("\n") if x.strip()
                ]:
                    if is_generic_email(email, blocked_emails, suspicious_email_patterns):
                        errors.append(f"Row {idx}: forbidden or suspicious Corporate email `{email}`")

            if mobile_phone:
                if has_suspicious_phone(mobile_phone, suspicious_phone_patterns):
                    errors.append(f"Row {idx}: suspicious Mobile phone `{mobile_phone}`")

            if work_phone_ext:
                if "доб" not in work_phone_ext.lower() and "ext" not in work_phone_ext.lower():
                    errors.append(f"Row {idx}: Work phone + extension missing named extension marker")
                if has_suspicious_phone(work_phone_ext, suspicious_phone_patterns):
                    errors.append(f"Row {idx}: suspicious Work phone + extension `{work_phone_ext}`")

            if telegram_nicknames and "http" in telegram_nicknames.lower():
                errors.append(f"Row {idx}: Telegram niknames should store handles, not URLs")
            if telegram_group and "http" in telegram_group.lower():
                errors.append(f"Row {idx}: Telegram-group should store handles, not URLs")

            if any([corporate_email, mobile_phone, work_phone_ext, telegram_nicknames]):
                contact_qualified += 1

        if contact_qualified < 50:
            warnings.append(
                f"Only {contact_qualified} rows currently count toward the 50-company direct-contact threshold"
            )

        if TARGETS.exists():
            with TARGETS.open(encoding="utf-8", newline="") as f:
                seed_count = sum(1 for _ in csv.DictReader(f))
            if ws.max_row - 1 < seed_count:
                warnings.append(
                    f"Workbook has {ws.max_row - 1} rows; seed list has {seed_count} companies"
                )

    if not EVIDENCE.exists():
        errors.append(f"Missing evidence log: {EVIDENCE}")
    else:
        with EVIDENCE.open(encoding="utf-8", newline="") as f:
            evidence_rows = list(csv.DictReader(f))
        if not evidence_rows:
            warnings.append("Evidence log is empty")

    if errors:
        print("VALIDATION FAILED")
        print("=" * 80)
        for item in errors:
            print("ERROR:", item)
        if warnings:
            print("-" * 80)
            for item in warnings:
                print("WARN:", item)
        return 1

    print("VALIDATION PASSED")
    if warnings:
        print("=" * 80)
        for item in warnings:
            print("WARN:", item)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
