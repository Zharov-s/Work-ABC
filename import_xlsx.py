#!/usr/bin/env python3
"""Import companies from Производство.xlsx into the contacts database.

Dedup rules:
- Skip if INN already exists in contacts
- Skip if normalized company name already exists
- UNIQUE(email) ON CONFLICT IGNORE handles email-level conflicts

Run with --backfill to update okved for already-imported records only.
"""
import re
import sqlite3
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl

XLSX_PATH = Path(__file__).parent / "Research/output/Производство.xlsx"
DB_PATH = Path(__file__).parent / "webapp/data/contacts.db"

GENERIC_PREFIXES = {
    "info", "mail", "sales", "office", "support", "contact", "admin",
    "help", "noreply", "no-reply", "orders", "zakupki", "press",
    "service", "biz", "manager", "reception", "sekretariat", "sekr",
    "hr", "marketing", "pr", "media", "tender", "tenders",
    "krasnodar", "kavkaz", "msk", "spb", "moscow", "nn", "ekb",
    "post", "email", "hello", "team", "general", "main", "central",
}


def normalize_name(name: str) -> str:
    name = name.lower().strip()
    name = unicodedata.normalize("NFKC", name)
    name = re.sub(r'["""\'«»\-–—]', "", name)
    name = re.sub(r'\s+', " ", name).strip()
    for prefix in ("ооо ", "зао ", "оао ", "пао ", "ао ", "ип ", "гк ", "нпо ", "нпп ", "спк "):
        if name.startswith(prefix):
            name = name[len(prefix):]
            break
    return name.strip()


def extract_okved_code(raw: str | None) -> str | None:
    """Extract OKVED code from string like '28.22 Производство подъёмного оборудования'."""
    if not raw:
        return None
    m = re.match(r'^(\d{2}(?:\.\d+)*)', str(raw).strip())
    return m.group(1) if m else None


def classify_emails(emails: list[str]) -> tuple[str | None, str | None, str | None]:
    """Returns (primary_email, personal_email, generic_email)."""
    personal = None
    generic = None
    fallback = None

    for e in emails:
        local = e.split("@")[0].lower()
        is_generic = local in GENERIC_PREFIXES or any(local.startswith(g) for g in GENERIC_PREFIXES)

        if "." in local and not is_generic:
            if personal is None:
                personal = e
        elif is_generic:
            if generic is None:
                generic = e
        else:
            if fallback is None:
                fallback = e

    primary = personal or fallback or generic
    return primary, personal, generic


def parse_phones(raw: str | None) -> tuple[str | None, str | None, str | None]:
    """Returns (phone, mobile_phone, generic_phone)."""
    if not raw:
        return None, None, None
    phones = [p.strip() for p in raw.split("\n") if p.strip()]
    phones = [p for p in phones if re.search(r'\d', p)]
    return (
        phones[0] if len(phones) > 0 else None,
        phones[1] if len(phones) > 1 else None,
        phones[2] if len(phones) > 2 else None,
    )


def parse_emails(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [e.strip().lower() for e in raw.split() if "@" in e.strip()]


def backfill_okved() -> None:
    """Update okved column for records already imported from Производство.xlsx."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    updated = 0
    for row in rows:
        inn = str(row[1]).strip() if row[1] else None
        okved = extract_okved_code(row[8])
        if not okved:
            continue

        if inn:
            r = conn.execute("SELECT id, okved FROM contacts WHERE inn=?", (inn,)).fetchone()
            if r and not r["okved"]:
                conn.execute("UPDATE contacts SET okved=? WHERE id=?", (okved, r["id"]))
                updated += 1
        else:
            company_name = (row[0] or "").strip()
            norm = normalize_name(company_name)
            rows_db = conn.execute(
                "SELECT id, okved, company_name FROM contacts WHERE source_url='Производство.xlsx'"
            ).fetchall()
            for db_row in rows_db:
                if normalize_name(db_row["company_name"]) == norm and not db_row["okved"]:
                    conn.execute("UPDATE contacts SET okved=? WHERE id=?", (okved, db_row["id"]))
                    updated += 1

    conn.commit()
    conn.close()
    print(f"Backfill complete: {updated} records updated with okved codes")


def main() -> None:
    if not XLSX_PATH.exists():
        print(f"ERROR: File not found: {XLSX_PATH}")
        sys.exit(1)
    if not DB_PATH.exists():
        print(f"ERROR: Database not found: {DB_PATH}")
        sys.exit(1)

    if "--backfill" in sys.argv:
        backfill_okved()
        return

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"Loaded {len(rows)} rows from xlsx")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    existing_inns = {
        r[0] for r in conn.execute("SELECT inn FROM contacts WHERE inn IS NOT NULL AND inn != ''").fetchall()
    }
    existing_names = {
        normalize_name(r[0])
        for r in conn.execute("SELECT company_name FROM contacts").fetchall()
    }
    existing_emails = {
        r[0].lower()
        for r in conn.execute("SELECT email FROM contacts WHERE email IS NOT NULL").fetchall()
    }

    today = datetime.now().strftime("%Y-%m-%d")
    inserted = 0
    skipped_inn = 0
    skipped_name = 0
    skipped_email = 0

    for row in rows:
        company_name = (row[0] or "").strip()
        inn = str(row[1]).strip() if row[1] else None
        person_name = (row[2] or "").strip() or None
        title = (row[4] or "").strip() or None
        phone_raw = row[5]
        email_raw = row[6]
        website = (row[7] or "").strip() or None
        okved = extract_okved_code(row[8])
        segment = (row[12] or "").strip() or None

        if not company_name:
            continue

        if inn and inn in existing_inns:
            skipped_inn += 1
            continue

        norm = normalize_name(company_name)
        if norm in existing_names:
            skipped_name += 1
            continue

        emails = parse_emails(email_raw)
        primary, personal_email, generic_email = classify_emails(emails)

        if primary and primary.lower() in existing_emails:
            skipped_email += 1
            continue

        phone, mobile_phone, generic_phone = parse_phones(phone_raw)

        conn.execute(
            """INSERT OR IGNORE INTO contacts
               (company_name, inn, person_name, title,
                email, personal_email, generic_email,
                phone, mobile_phone, generic_phone,
                website, segment, okved, date_found, source_url, status)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                company_name, inn, person_name, title,
                primary, personal_email, generic_email,
                phone, mobile_phone, generic_phone,
                website, segment, okved, today, "Производство.xlsx", "new",
            )
        )
        inserted += 1
        existing_names.add(norm)
        if inn:
            existing_inns.add(inn)
        if primary:
            existing_emails.add(primary.lower())

    conn.commit()
    conn.close()

    print(f"\nImport complete:")
    print(f"  Inserted:        {inserted}")
    print(f"  Skipped by INN:  {skipped_inn}")
    print(f"  Skipped by name: {skipped_name}")
    print(f"  Skipped by email:{skipped_email}")
    print(f"  Total skipped:   {skipped_inn + skipped_name + skipped_email}")


if __name__ == "__main__":
    main()
